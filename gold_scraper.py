#!/usr/bin/env python3
"""
Gold Price Scraper for qatar-goldprice.com
Extracts gold prices for different karats (14, 18, 21, 22, 24)
"""

import requests
from bs4 import BeautifulSoup
import json
import re
from datetime import datetime, timedelta
from typing import Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class GoldPriceScraper:
    """Scraper for gold prices from qatar-goldprice.com"""
    
    def __init__(self):
        self.base_url = "https://qatar-goldprice.com/"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.target_karats = [14, 18, 21, 22, 24]
    
    def fetch_page(self) -> Optional[BeautifulSoup]:
        """Fetch the webpage and return BeautifulSoup object"""
        try:
            response = requests.get(self.base_url, headers=self.headers, timeout=10)
            response.raise_for_status()
            response.encoding = 'utf-8'
            return BeautifulSoup(response.text, 'html.parser')
        except requests.RequestException as e:
            print(f"Error fetching page: {e}")
            return None
    
    def extract_gold_prices(self, soup: BeautifulSoup) -> Dict:
        """Extract gold prices from the webpage"""
        prices = {
            'timestamp': datetime.now().isoformat(),
            'source': self.base_url,
            'prices': {}
        }
        
        # Find all tables on the page
        tables = soup.find_all('table')
        
        for table in tables:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 3:
                    # Get text from all cells
                    cell_texts = [cell.get_text(strip=True) for cell in cells]
                    row_text = ' '.join(cell_texts)
                    
                    # Check for each target karat in Arabic format
                    for karat in self.target_karats:
                        # Look for Arabic text pattern: "عيار {karat}" or just the karat number
                        karat_patterns = [
                            f'عيار {karat}',
                            f'جرام الذهب عيار {karat}',
                            f'{karat}'
                        ]
                        
                        found_karat = False
                        for pattern in karat_patterns:
                            if pattern in row_text:
                                found_karat = True
                                break
                        
                        if found_karat:
                            # Extract prices from the row
                            # Typically: [description, QAR price, USD price]
                            qar_price = None
                            usd_price = None
                            
                            # Look for numeric values in cells (skip first cell which is description)
                            numeric_cells = []
                            for cell in cells[1:]:
                                cell_text = cell.get_text(strip=True)
                                # Remove commas and try to parse as float
                                cell_text_clean = cell_text.replace(',', '').replace(' ', '')
                                try:
                                    price = float(cell_text_clean)
                                    numeric_cells.append(price)
                                except ValueError:
                                    continue
                            
                            # First numeric value is usually QAR, second is USD
                            if len(numeric_cells) >= 1:
                                qar_price = numeric_cells[0]
                            if len(numeric_cells) >= 2:
                                usd_price = numeric_cells[1]
                            
                            if qar_price is not None:
                                prices['prices'][f'{karat}k'] = {
                                    'karat': karat,
                                    'price_qar': qar_price,
                                    'price_usd': usd_price,
                                    'unit': 'gram'
                                }
                                break  # Found this karat, move to next row
        
        # If no prices found in tables, try alternative method using regex on page text
        if not prices['prices']:
            page_text = soup.get_text()
            
            for karat in self.target_karats:
                # Look for pattern: "عيار {karat}" followed by numbers
                pattern = rf'عيار\s*{karat}.*?(\d+[.,]\d+).*?(\d+[.,]\d+)'
                matches = re.findall(pattern, page_text, re.DOTALL)
                
                if matches:
                    try:
                        qar_price = float(matches[0][0].replace(',', ''))
                        usd_price = float(matches[0][1].replace(',', ''))
                        prices['prices'][f'{karat}k'] = {
                            'karat': karat,
                            'price_qar': qar_price,
                            'price_usd': usd_price,
                            'unit': 'gram'
                        }
                    except (ValueError, IndexError):
                        continue
        
        return prices
    
    def scrape(self) -> Dict:
        """Main method to scrape gold prices"""
        soup = self.fetch_page()
        if not soup:
            return {'error': 'Failed to fetch webpage'}
        
        prices = self.extract_gold_prices(soup)
        return prices
    
    def print_prices(self, prices: Dict):
        """Print prices in a formatted way"""
        if 'error' in prices:
            print(f"Error: {prices['error']}")
            return
        
        print("\n" + "="*60)
        print("GOLD PRICES FROM QATAR-GOLDPRICE.COM")
        print("="*60)
        print(f"Timestamp: {prices.get('timestamp', 'N/A')}")
        print(f"Source: {prices.get('source', 'N/A')}")
        print("\nPrice per Gram:")
        print("-"*60)
        
        if prices.get('prices'):
            for karat_key in sorted(prices['prices'].keys(), key=lambda x: int(x.replace('k', ''))):
                price_info = prices['prices'][karat_key]
                print(f"  {price_info['karat']}K Gold:")
                if price_info.get('price_qar'):
                    print(f"    QAR: {price_info['price_qar']:,.2f}")
                if price_info.get('price_usd'):
                    print(f"    USD: {price_info['price_usd']:,.2f}")
                print()
        else:
            print("  No prices found. The website structure may have changed.")
        
        print("="*60)
    
    def export_to_excel(self, prices: Dict, filename: str = 'gold_prices.xlsx'):
        """Export gold prices to Excel file in RTL (right-to-left) table format"""
        if 'error' in prices:
            print(f"Error: Cannot export - {prices['error']}")
            return False
        
        # Create or load workbook
        karat_col = 1  # Column A contains karat labels (appears rightmost in RTL view)
        day_name_row = 1  # Row 1: Arabic day names
        date_row = 2  # Row 2: Day number and month, and "نوع العيار" header
        
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            # Ensure RTL is set
            ws.sheet_view.rightToLeft = True
            # Find the last date column (columns after A are dates)
            max_col = ws.max_column
            if max_col < 1:
                max_col = 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Gold Prices"
            
            # Set sheet to RTL (right-to-left) direction
            ws.sheet_view.rightToLeft = True
            
            # Initialize karat labels in column A (rightmost in RTL view)
            # Rows 3-7 for karat labels (14, 18, 21, 22, 24)
            karat_row_map = {14: 3, 18: 4, 21: 5, 22: 6, 24: 7}
            karat_labels = {14: '14', 18: '18', 21: '21', 22: '22', 24: '24'}
            
            # Place karat labels in column A
            for karat, row_num in karat_row_map.items():
                label_cell = ws.cell(row=row_num, column=karat_col)
                label_cell.value = karat_labels[karat]
                label_cell.font = Font(bold=True)
                label_cell.alignment = Alignment(horizontal='center', vertical='center')
                # Light orange/peach background for karat labels
                label_cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
                # Special styling for 22K (red text)
                if karat == 22:
                    label_cell.font = Font(bold=True, color='FF0000')
            
            # Add "نوع العيار" header in row 2, column A (same row as date headers)
            header_label_cell = ws.cell(row=date_row, column=karat_col)
            header_label_cell.value = 'نوع العيار'  # Type of Karat
            header_label_cell.font = Font(bold=True)
            header_label_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
            header_label_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            max_col = 1  # Start with one column (karat labels)
        
        # Get today's date
        today = datetime.now()
        date_str = today.strftime('%Y-%m-%d')
        day_name_ar = self._get_arabic_day_name(today.weekday())
        month_name_ar = self._get_arabic_month_name(today.month)
        day_name_text = day_name_ar  # Row 1: Just the day name
        date_text = f'{month_name_ar} {today.day}'  # Row 2: Month and day number
        
        # Check if today's date column already exists
        # Dates are in columns B, C, D... (after column A which has karat labels)
        date_col = None
        for col in range(2, max_col + 2):  # Start from column B (skip column A)
            # Check row 2 (date row) for matching date
            cell_value = ws.cell(row=date_row, column=col).value
            if cell_value and (date_str in str(cell_value) or date_text in str(cell_value)):
                date_col = col
                break
        
        # If date doesn't exist, add new column after the last date column
        if date_col is None:
            # Find the last date column (last column that's not A)
            last_date_col = max_col
            if last_date_col < 2:
                last_date_col = 1
            date_col = last_date_col + 1
        
        # Set row 1: Arabic day name
        day_name_cell = ws.cell(row=day_name_row, column=date_col)
        day_name_cell.value = day_name_text
        day_name_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        day_name_cell.font = Font(bold=True)
        day_name_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set row 2: Month and day number
        date_cell = ws.cell(row=date_row, column=date_col)
        date_cell.value = date_text
        date_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ensure karat labels are in column A (rows 3-7)
        karat_row_map = {14: 3, 18: 4, 21: 5, 22: 6, 24: 7}
        karat_labels = {14: '14', 18: '18', 21: '21', 22: '22', 24: '24'}
        
        for karat, row_num in karat_row_map.items():
            label_cell = ws.cell(row=row_num, column=karat_col)
            if label_cell.value is None:
                label_cell.value = karat_labels[karat]
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            label_cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
            if karat == 22:
                label_cell.font = Font(bold=True, color='FF0000')
        
        # Ensure "نوع العيار" header is in column A, row 2 (same row as date headers)
        header_label_cell = ws.cell(row=date_row, column=karat_col)
        if header_label_cell.value is None or header_label_cell.value != 'نوع العيار':
            header_label_cell.value = 'نوع العيار'
        header_label_cell.font = Font(bold=True)
        header_label_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ensure we have enough rows
        for row_num in range(3, 8):
            if ws.max_row < row_num:
                ws.append([None] * ws.max_column)
        
        # Write prices to appropriate rows (rows 3-7)
        for karat_key, price_info in prices.get('prices', {}).items():
            karat = price_info['karat']
            if karat in karat_row_map:
                row_num = karat_row_map[karat]
                # Use USD price (or QAR if USD not available)
                price = price_info.get('price_usd') or price_info.get('price_qar', 0)
                price_cell = ws.cell(row=row_num, column=date_col)
                price_cell.value = price
                price_cell.number_format = '0.00'
                price_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 12  # Karat label column
        for col in range(2, date_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Ensure RTL is set
        ws.sheet_view.rightToLeft = True
        
        # Save workbook
        wb.save(filename)
        print(f"\nPrices exported to {filename} (RTL layout)")
        return True
    
    def _get_arabic_day_name(self, weekday: int) -> str:
        """Get Arabic day name"""
        days = {
            0: 'الاثنين',  # Monday
            1: 'الثلاثاء',  # Tuesday
            2: 'الأربعاء',  # Wednesday
            3: 'الخميس',    # Thursday
            4: 'الجمعة',    # Friday
            5: 'السبت',     # Saturday
            6: 'الأحد'      # Sunday
        }
        return days.get(weekday, '')
    
    def _get_arabic_month_name(self, month: int) -> str:
        """Get Arabic month name"""
        months = {
            1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'أبريل',
            5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
            9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
        }
        return months.get(month, '')


def main():
    """Main function"""
    scraper = GoldPriceScraper()
    prices = scraper.scrape()
    scraper.print_prices(prices)
    
    # Save to JSON
    json_file = 'gold_prices.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(prices, f, indent=2, ensure_ascii=False)
    print(f"\nPrices saved to {json_file}")
    
    # Export to Excel
    excel_file = 'gold_prices.xlsx'
    scraper.export_to_excel(prices, excel_file)


if __name__ == "__main__":
    main()
