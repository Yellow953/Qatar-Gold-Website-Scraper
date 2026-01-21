#!/usr/bin/env python3
"""
Hotel Price Scraper for booking.com
Extracts hotel prices for hotels in Doha, Qatar
"""

try:
    import undetected_chromedriver as uc
    USE_UNDETECTED = True
except ImportError:
    USE_UNDETECTED = False

import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import json
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


class HotelPriceScraper:
    """Scraper for hotel prices from booking.com"""
    
    def __init__(self, headless=False):
        self.base_url = "https://www.booking.com"
        self.hotels = self._get_hotel_list()
        self.driver = None
        self.headless = headless
        
    def _get_hotel_list(self) -> List[str]:
        """Get list of hotels to scrape"""
        hotels = [
            "فندق فور سيزونز الدوحة",
            "فندق جراند حياة الدوحة",
            "فندق إنتركونتيننتال الدوحة",
            "فندق ماريوت الدوحة",
            "فندق ومنتجع شيراتون الدوحة",
            "فندق كونكورد الدوحة",
            "فندق جراند ميركيور سيتى سنتر الدوحة",
            "فندق رويال قطر",
            "فندق ريتاج الريان",
            "فندق راديسون بلو بروت مان",
            "فندق موفنبيك الدوحة",
            "فنادق إزدان الدوحة",
            "فندق لي بارك",
            "فندق و اجنحة سراي مشيرب",
            "اجنحة الليوان الفندقية",
            "المنصور سويت هوتيل",
            "Gulf Pearl Hotel Apartments",
            "Mathema Premium Aparthotel",
            "فندق البستان",
            "فندق المنتزه بلازا",
            "فندق مشيرب",
            "فندق جراند قطر بالاس",
            "فندق جراند سويت",
            "فندق قصر لا فيلا",
            "رتاج ريزيدنس السد",
            "الصفا للاجنحة الملكية",
            "ازدان للاجنحة الفندقية",
            "أجنحة المدينة الدوحة",
            "المنصور بارك إن فندق وشقق فندقية",
            "فندق السد سويتس",
            "وايت مون ريزيدنس",
            "TGI Residence",
            "حياة ريزدنسز دوحة ويست باي"
        ]
        return hotels
    
    def _get_hotel_english_name(self, arabic_name: str) -> str:
        """Get English name for Arabic hotel name"""
        hotel_mapping = {
            "فندق فور سيزونز الدوحة": "Four Seasons Hotel Doha",
            "فندق جراند حياة الدوحة": "Grand Hyatt Doha",
            "فندق إنتركونتيننتال الدوحة": "InterContinental Doha",
            "فندق ماريوت الدوحة": "Marriott Doha",
            "فندق ومنتجع شيراتون الدوحة": "Sheraton Grand Doha Resort",
            "فندق كونكورد الدوحة": "Concorde Hotel Doha",
            "فندق جراند ميركيور سيتى سنتر الدوحة": "Grand Mercure Doha City Centre",
            "فندق رويال قطر": "Royal Qatar Hotel",
            "فندق ريتاج الريان": "Retaj Al Rayan Hotel",
            "فندق راديسون بلو بروت مان": "Radisson Blu Hotel Doha",
            "فندق موفنبيك الدوحة": "Mövenpick Hotel Doha",
            "فنادق إزدان الدوحة": "Ezdan Hotels Doha",
            "فندق لي بارك": "Le Park Hotel",
            "فندق و اجنحة سراي مشيرب": "Saray Msheireb Hotel",
            "اجنحة الليوان الفندقية": "Liwan Hotel Suites",
            "المنصور سويت هوتيل": "Al Mansour Suites Hotel",
            "فندق البستان": "Al Bustan Hotel",
            "فندق المنتزه بلازا": "Al Muntazah Plaza Hotel",
            "فندق مشيرب": "Msheireb Hotel",
            "فندق جراند قطر بالاس": "Grand Qatar Palace Hotel",
            "فندق جراند سويت": "Grand Suite Hotel",
            "فندق قصر لا فيلا": "La Villa Palace Hotel",
            "رتاج ريزيدنس السد": "Retaj Residence Al Sadd",
            "الصفا للاجنحة الملكية": "Al Safa Royal Suites",
            "ازدان للاجنحة الفندقية": "Ezdan Hotel Suites",
            "أجنحة المدينة الدوحة": "Madinat Doha Suites",
            "المنصور بارك إن فندق وشقق فندقية": "Park Inn by Radisson Al Mansour",
            "فندق السد سويتس": "Al Sadd Suites Hotel",
            "وايت مون ريزيدنس": "White Moon Residence",
            "حياة ريزدنسز دوحة ويست باي": "Hyatt Residences Doha West Bay"
        }
        return hotel_mapping.get(arabic_name, arabic_name)
    
    def _setup_driver(self, headless=False):
        """Setup Chrome WebDriver using undetected-chromedriver to avoid detection"""
        if USE_UNDETECTED:
            try:
                # Use undetected-chromedriver which is specifically designed to bypass detection
                options = uc.ChromeOptions()
                if headless:
                    options.add_argument('--headless=new')  # Use new headless mode
                
                # Additional options
                options.add_argument('--no-sandbox')
                options.add_argument('--disable-dev-shm-usage')
                options.add_argument('--start-maximized')
                
                # Create undetected Chrome driver
                self.driver = uc.Chrome(options=options, version_main=None)
                
                # Set window size to look more realistic
                if not headless:
                    self.driver.set_window_size(1920, 1080)
                
                print("    Using undetected-chromedriver to avoid detection")
                return
            except Exception as e:
                print(f"    Warning: Could not use undetected-chromedriver: {e}")
                print("    Falling back to standard Selenium")
        
        # Fallback to standard Selenium
        chrome_options = Options()
        if headless:
            chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
        # Execute scripts to avoid detection
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        self.driver.set_window_size(1920, 1080)
    
    def _close_driver(self):
        """Close the WebDriver"""
        if self.driver:
            self.driver.quit()
    
    def _search_hotel(self, hotel_name: str) -> Optional[Dict]:
        """Search for a hotel by interacting with booking.com like a human"""
        try:
            # Use 5 days in advance for check-in date, day after for checkout
            checkin_date = (datetime.now() + timedelta(days=1))  # Tomorrow
            checkout_date = (datetime.now() + timedelta(days=2))  # Day after tomorrow
            checkin_str = checkin_date.strftime('%Y-%m-%d')
            checkout_str = checkout_date.strftime('%Y-%m-%d')
            
            # Get English name if available
            english_name = self._get_hotel_english_name(hotel_name)
            search_query = english_name if english_name != hotel_name else hotel_name
            
            print(f"    Searching for: {search_query}")
            print(f"    Dates: {checkin_str} to {checkout_str}")
            
            # Navigate to booking.com homepage
            self.driver.get(self.base_url)
            time.sleep(5)
            
            # Close any consent/cookie dialogs
            try:
                consent_selectors = [
                    "button#onetrust-accept-btn-handler",
                    "button[id*='accept']",
                    "button[class*='accept']",
                    "[data-testid='cookie-consent-accept']",
                    "button[aria-label*='Accept']"
                ]
                for selector in consent_selectors:
                    try:
                        consent_btn = WebDriverWait(self.driver, 3).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                        )
                        consent_btn.click()
                        time.sleep(2)
                        break
                    except:
                        continue
            except:
                pass
            
            # Close sign-in discount popup if it appears
            try:
                # Look for the close button in the sign-in popup
                signin_popup_close = self.driver.find_elements(By.CSS_SELECTOR, 
                    "button[aria-label='Dismiss sign-in info.'], "
                    "button[aria-label*='Dismiss'], "
                    "[role='dialog'][aria-label*='sign in'] button[aria-label*='Dismiss'], "
                    "[role='dialog'] button[aria-label*='Dismiss sign-in']"
                )
                if signin_popup_close:
                    signin_popup_close[0].click()
                    print("    Closed sign-in popup")
                    time.sleep(2)
            except:
                pass
            
            # Find and fill the search box
            try:
                # Multiple selectors for search box
                search_selectors = [
                    "input[name='ss']",
                    "input[placeholder*='Where are you going']",
                    "input[data-testid='searchbox-destination-input']",
                    "#ss"
                ]
                
                search_box = None
                for selector in search_selectors:
                    try:
                        search_box = WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        if search_box:
                            break
                    except:
                        continue
                
                if not search_box:
                    print("    Could not find search box")
                    return None
                
                # Clear and type hotel name
                search_box.clear()
                time.sleep(1)
                search_box.send_keys(search_query)
                time.sleep(3)  # Wait for autocomplete suggestions
                
                # Try to select from autocomplete if available
                try:
                    # Wait for suggestions and click first one
                    suggestion = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "[data-testid='autocomplete-result'], .c-autocomplete__item, .sb-autocomplete__item"))
                    )
                    suggestion.click()
                    time.sleep(2)
                except:
                    # If no suggestion, just continue with typed text
                    pass
                
            except Exception as e:
                print(f"    Error filling search box: {e}")
                return None
            
            # Select check-in date
            try:
                # Find and click date picker to open calendar
                date_selectors = [
                    "[data-testid='date-display-field-start']",
                    "[data-testid='searchbox-datepicker-start']",
                    ".sb-date-field__display",
                    "button[data-testid='date-display-field-start']",
                    "[data-testid='date-display-field-end']",
                    ".sb-date-field"
                ]
                
                date_picker = None
                for selector in date_selectors:
                    try:
                        date_picker = self.driver.find_element(By.CSS_SELECTOR, selector)
                        if date_picker:
                            date_picker.click()
                            print("    Opened date picker")
                            time.sleep(3)
                            break
                    except:
                        continue
                
                if not date_picker:
                    print("    Warning: Could not open date picker")
                else:
                    # Navigate to correct month if needed
                    try:
                        # Get target month and year
                        target_month = checkin_date.strftime('%B %Y')  # e.g., "January 2026"
                        target_month_short = checkin_date.strftime('%b %Y')  # e.g., "Jan 2026"
                        
                        # Wait for calendar to load
                        time.sleep(2)
                        
                        # First, check if the target date is already visible (no need to navigate)
                        checkin_date_visible = False
                        try:
                            WebDriverWait(self.driver, 5).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "span[data-date]"))
                            )
                            all_date_spans = self.driver.find_elements(By.CSS_SELECTOR, f"span[data-date='{checkin_str}']")
                            if all_date_spans:
                                for span in all_date_spans:
                                    try:
                                        aria_disabled = span.get_attribute('aria-disabled')
                                        class_attr = span.get_attribute('class') or ''
                                        if aria_disabled != 'true' and 'ad9d5181d0' not in class_attr:
                                            checkin_date_visible = True
                                            break
                                    except:
                                        pass
                        except:
                            pass
                        
                        # Only navigate if the target date is not visible
                        if not checkin_date_visible:
                            # Check current month displayed - look for month name in calendar
                            # Try multiple selectors to find the actual month name
                            current_month_elem = []
                            month_selectors = [
                                "h3[aria-live='polite']",
                                ".bui-calendar__month",
                                ".bui-calendar__display-month",
                                "[data-testid='calendar-month']",
                                "h3.af236b7586",
                                ".e7addce19e",
                                "h3.bui-calendar__month",
                                "h3"  # Fallback: all h3 elements
                            ]
                            
                            month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                                         'July', 'August', 'September', 'October', 'November', 'December',
                                         'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                         'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                            
                            for selector in month_selectors:
                                try:
                                    elems = self.driver.find_elements(By.CSS_SELECTOR, selector)
                                    for elem in elems:
                                        text = elem.text.strip()
                                        # Check if it looks like a month name (contains month name and year)
                                        if any(month in text for month in month_names) and any(char.isdigit() for char in text):
                                            current_month_elem = [elem]
                                            break
                                    if current_month_elem:
                                        break
                                except:
                                    continue
                            # Check if we're already on the target month before navigating
                            target_month_num = checkin_date.month
                            target_year = checkin_date.year
                            navigated = False
                            
                            # First, try to detect current month
                            month_text = ""
                            current_month_num = None
                            current_year = None
                            
                            try:
                                if current_month_elem:
                                    month_text = current_month_elem[0].text.strip()
                                    print(f"    Current month displayed: {month_text}")
                                    
                                    # Parse month and year from text
                                    for idx, month_name in enumerate(['January', 'February', 'March', 'April', 'May', 'June', 
                                                                    'July', 'August', 'September', 'October', 'November', 'December'], 1):
                                        if month_name.lower() in month_text.lower():
                                            current_month_num = idx
                                            # Extract year (4 digits)
                                            year_match = re.search(r'\b(20\d{2})\b', month_text)
                                            if year_match:
                                                current_year = int(year_match.group(1))
                                            break
                            except:
                                pass
                            
                            # Check if we're already on the target month
                            if current_month_num == target_month_num and current_year == target_year:
                                print(f"    Already on target month: {target_month}")
                                navigated = True
                            else:
                                # Only navigate if we're not on the target month
                                print(f"    Need to navigate: current={current_month_num}/{current_year}, target={target_month_num}/{target_year}")
                            
                            max_navigations = 12
                            for i in range(max_navigations):
                                # Re-check current month
                                try:
                                    if current_month_elem:
                                        month_text = current_month_elem[0].text.strip()
                                        # Re-parse
                                        current_month_num = None
                                        current_year = None
                                        for idx, month_name in enumerate(['January', 'February', 'March', 'April', 'May', 'June', 
                                                                        'July', 'August', 'September', 'October', 'November', 'December'], 1):
                                            if month_name.lower() in month_text.lower():
                                                current_month_num = idx
                                                year_match = re.search(r'\b(20\d{2})\b', month_text)
                                                if year_match:
                                                    current_year = int(year_match.group(1))
                                                break
                                except:
                                    pass
                                
                                # Check if we're now on target
                                if current_month_num == target_month_num and current_year == target_year:
                                    print(f"    Found target month: {target_month}")
                                    navigated = True
                                    break
                                
                                # Determine navigation direction
                                need_forward = False
                                need_backward = False
                                
                                if current_month_num and current_year:
                                    # Compare dates
                                    if current_year < target_year or (current_year == target_year and current_month_num < target_month_num):
                                        need_forward = True
                                    elif current_year > target_year or (current_year == target_year and current_month_num > target_month_num):
                                        need_backward = True
                                
                                # Click appropriate button
                                try:
                                    if need_forward:
                                        button = self.driver.find_element(By.CSS_SELECTOR, 
                                            "button[aria-label='Next month'], "
                                            "button[aria-label*='Next'], "
                                            ".bui-calendar__control--next, "
                                            "[data-testid='calendar-next-month']"
                                        )
                                        direction = "forward"
                                    elif need_backward:
                                        button = self.driver.find_element(By.CSS_SELECTOR, 
                                            "button[aria-label='Previous month'], "
                                            "button[aria-label*='Previous'], "
                                            ".bui-calendar__control--prev, "
                                            "[data-testid='calendar-prev-month']"
                                        )
                                        direction = "backward"
                                    else:
                                        # Can't determine direction, skip navigation
                                        print(f"    Cannot determine navigation direction, skipping")
                                        break
                                    
                                    button.click()
                                    print(f"    Clicked {direction} month button (attempt {i+1})")
                                    time.sleep(2)
                                    
                                    # Re-find month element after navigation
                                    current_month_elem = []
                                    for selector in month_selectors:
                                        try:
                                            elems = self.driver.find_elements(By.CSS_SELECTOR, selector)
                                            for elem in elems:
                                                text = elem.text.strip()
                                                if any(month in text for month in month_names) and any(char.isdigit() for char in text):
                                                    current_month_elem = [elem]
                                                    break
                                            if current_month_elem:
                                                break
                                        except:
                                            continue
                                except Exception as e:
                                    print(f"    Could not find/click navigation button: {e}")
                                    break
                            
                            if not navigated:
                                print(f"    Warning: Could not navigate to target month {target_month}, trying to select date anyway")
                    except Exception as e:
                        print(f"    Warning: Could not navigate to target month: {e}")
                    else:
                        print(f"    Target date {checkin_str} is already visible, skipping navigation")
                    
                    # Wait for calendar to fully load
                    time.sleep(1)
                    
                    # Find and click the check-in date
                    # Based on HTML: <span class="ecb788f3b7 c0b8f1e8f8" data-date="2026-01-23" ...>
                    date_found = False
                    
                    try:
                        # Wait for calendar dates to be present
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "span[data-date]"))
                        )
                        time.sleep(1)  # Additional wait for calendar to fully render
                        
                        # Find all spans with data-date attribute matching our date
                        all_date_spans = self.driver.find_elements(By.CSS_SELECTOR, f"span[data-date='{checkin_str}']")
                        
                        print(f"    Found {len(all_date_spans)} span(s) with data-date='{checkin_str}'")
                        
                        if all_date_spans:
                            for span in all_date_spans:
                                try:
                                    # Check if it's not disabled
                                    aria_disabled = span.get_attribute('aria-disabled')
                                    class_attr = span.get_attribute('class') or ''
                                    
                                    print(f"    Checking span: aria-disabled={aria_disabled}, class={class_attr[:50]}")
                                    
                                    if aria_disabled != 'true' and 'ad9d5181d0' not in class_attr:
                                        # Scroll into view if needed
                                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", span)
                                        time.sleep(0.5)
                                        
                                        # Try clicking with JavaScript if regular click doesn't work
                                        try:
                                            span.click()
                                        except:
                                            self.driver.execute_script("arguments[0].click();", span)
                                        
                                        print(f"    Selected check-in date: {checkin_str}")
                                        date_found = True
                                        time.sleep(2)
                                        break
                                except Exception as e:
                                    print(f"    Error clicking span: {e}")
                                    continue
                        
                        if not date_found:
                            print(f"    Check-in date {checkin_str} not found or is disabled")
                            # Debug: show what dates are available
                            try:
                                all_dates = self.driver.find_elements(By.CSS_SELECTOR, "span[data-date]")
                                print(f"    Total date spans found: {len(all_dates)}")
                                if all_dates:
                                    sample_dates = [d.get_attribute('data-date') for d in all_dates[:10]]
                                    print(f"    Sample dates found: {sample_dates}")
                            except:
                                pass
                    except Exception as e:
                        print(f"    Error selecting check-in date {checkin_str}: {e}")
            except Exception as e:
                print(f"    Warning: Error setting check-in date: {e}")
            
            # Select check-out date
            try:
                date_found = False
                
                try:
                    # Wait a bit after check-in selection for calendar to update
                    time.sleep(2)
                    
                    # Check if calendar is still open, if not try to find check-out date picker
                    calendar_visible = False
                    try:
                        calendar_elem = self.driver.find_element(By.CSS_SELECTOR, 
                            ".bui-calendar, [data-testid='datepicker'], .sb-date-picker"
                        )
                        if calendar_elem.is_displayed():
                            calendar_visible = True
                    except:
                        pass
                    
                    # If calendar closed, try clicking on check-out date field to reopen
                    if not calendar_visible:
                        print("    Calendar closed, trying to reopen for check-out date")
                        checkout_date_selectors = [
                            "[data-testid='date-display-field-checkout']",
                            ".sb-date-field__display--checkout",
                            "input[name='checkout']",
                            ".checkout-date"
                        ]
                        for selector in checkout_date_selectors:
                            try:
                                checkout_field = self.driver.find_element(By.CSS_SELECTOR, selector)
                                checkout_field.click()
                                time.sleep(2)
                                print("    Reopened calendar for check-out")
                                break
                            except:
                                continue
                    
                    # Navigate to correct month for check-out if needed
                    try:
                        target_month = checkout_date.strftime('%B %Y')
                        target_month_short = checkout_date.strftime('%b %Y')
                        
                        # Find month element
                        current_month_elem = []
                        month_selectors = [
                            "h3[aria-live='polite']",
                            ".bui-calendar__month",
                            ".bui-calendar__display-month",
                            "[data-testid='calendar-month']",
                            "h3.af236b7586",
                            "h3"
                        ]
                        
                        month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                                     'July', 'August', 'September', 'October', 'November', 'December',
                                     'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                     'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                        
                        for selector in month_selectors:
                            try:
                                elems = self.driver.find_elements(By.CSS_SELECTOR, selector)
                                for elem in elems:
                                    text = elem.text.strip()
                                    if any(month in text for month in month_names) and any(char.isdigit() for char in text):
                                        current_month_elem = [elem]
                                        break
                                if current_month_elem:
                                    break
                            except:
                                continue
                        
                        # Navigate to target month if needed
                        if current_month_elem:
                            month_text = current_month_elem[0].text.strip()
                            if target_month.lower() not in month_text.lower() and target_month_short.lower() not in month_text.lower():
                                # Need to navigate
                                max_nav = 12
                                for i in range(max_nav):
                                    try:
                                        next_button = self.driver.find_element(By.CSS_SELECTOR, 
                                            "button[aria-label='Next month'], "
                                            "button[aria-label*='Next'], "
                                            ".bui-calendar__control--next"
                                        )
                                        next_button.click()
                                        time.sleep(1)
                                        # Re-check month
                                        for selector in month_selectors:
                                            try:
                                                elems = self.driver.find_elements(By.CSS_SELECTOR, selector)
                                                for elem in elems:
                                                    text = elem.text.strip()
                                                    if any(month in text for month in month_names) and any(char.isdigit() for char in text):
                                                        month_text = text
                                                        break
                                                if target_month.lower() in month_text.lower() or target_month_short.lower() in month_text.lower():
                                                    break
                                            except:
                                                continue
                                        if target_month.lower() in month_text.lower() or target_month_short.lower() in month_text.lower():
                                            break
                                    except:
                                        break
                    except Exception as e:
                        print(f"    Warning: Could not navigate to check-out month: {e}")
                    
                    # Wait for calendar dates to load
                    time.sleep(1)
                    
                    # Find all spans with data-date attribute for checkout
                    all_date_spans = self.driver.find_elements(By.CSS_SELECTOR, f"span[data-date='{checkout_str}']")
                    
                    print(f"    Found {len(all_date_spans)} span(s) with data-date='{checkout_str}'")
                    
                    if all_date_spans:
                        for span in all_date_spans:
                            try:
                                # Check if it's not disabled
                                aria_disabled = span.get_attribute('aria-disabled')
                                class_attr = span.get_attribute('class') or ''
                                
                                print(f"    Checking span: aria-disabled={aria_disabled}, class={class_attr[:50]}")
                                
                                # Available dates have aria-disabled != 'true' and don't have 'ad9d5181d0' class
                                if aria_disabled != 'true' and 'ad9d5181d0' not in class_attr:
                                    # Scroll into view if needed
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", span)
                                    time.sleep(0.5)
                                    
                                    # Try clicking with JavaScript if regular click doesn't work
                                    try:
                                        span.click()
                                    except:
                                        self.driver.execute_script("arguments[0].click();", span)
                                    
                                    print(f"    Selected check-out date: {checkout_str}")
                                    date_found = True
                                    time.sleep(2)
                                    break
                            except Exception as e:
                                print(f"    Error clicking span: {e}")
                                continue
                    
                    if not date_found:
                        print(f"    Check-out date {checkout_str} not found or is disabled")
                        # Debug: show what dates are available
                        try:
                            all_dates = self.driver.find_elements(By.CSS_SELECTOR, "span[data-date]")
                            print(f"    Total date spans found: {len(all_dates)}")
                            if all_dates:
                                sample_dates = [d.get_attribute('data-date') for d in all_dates[:10]]
                                print(f"    Sample dates found: {sample_dates}")
                        except:
                            pass
                except Exception as e:
                    print(f"    Error selecting check-out date {checkout_str}: {e}")
            except Exception as e:
                print(f"    Warning: Error setting check-out date: {e}")
            
            # Click search button
            try:
                search_button_selectors = [
                    "button[type='submit']",
                    "[data-testid='searchbox-submit-button']",
                    ".sb-searchbox__button",
                    "button.sb-searchbox__button"
                ]
                
                search_button = None
                for selector in search_button_selectors:
                    try:
                        search_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                        if search_button:
                            break
                    except:
                        continue
                
                if search_button:
                    search_button.click()
                    print("    Clicked search button")
                else:
                    # Fallback: press Enter on search box
                    search_box.send_keys("\n")
            except Exception as e:
                print(f"    Error clicking search: {e}")
                # Try pressing Enter as fallback
                try:
                    search_box.send_keys("\n")
                except:
                    pass
            
            # Wait for results page to load
            time.sleep(10)
            
            # Close sign-in popup if it appears on results page
            try:
                signin_popup_close = self.driver.find_elements(By.CSS_SELECTOR, 
                    "button[aria-label='Dismiss sign-in info.'], "
                    "button[aria-label*='Dismiss sign-in'], "
                    "[role='dialog'][aria-label*='sign in'] button[aria-label*='Dismiss'], "
                    "[role='dialog'][aria-label*='Window offering discounts'] button"
                )
                if signin_popup_close:
                    signin_popup_close[0].click()
                    print("    Closed sign-in popup on results page")
                    time.sleep(2)
            except:
                pass
            
            # Scroll a bit to trigger lazy loading
            self.driver.execute_script("window.scrollTo(0, 300);")
            time.sleep(3)
            
            # Check current URL
            current_url = self.driver.current_url
            print(f"    Current URL: {current_url[:100]}...")
            
            # Check if we're on a hotel page (not search results)
            if '/hotel/' in current_url and '/searchresults' not in current_url:
                # We were redirected to a hotel page - this is good!
                print("    Redirected to hotel page directly")
                return self._extract_price_from_hotel_page(hotel_name, search_query, current_url)
            
            # Try to find hotel in search results
            return self._extract_price_from_search_results(hotel_name, search_query)
            
        except Exception as e:
            print(f"  Error searching for {hotel_name}: {e}")
            return {
                'hotel_name': hotel_name,
                'price': None,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            
            # Try to find hotel in search results
            return self._extract_price_from_search_results(hotel_name, search_query)
            
        except Exception as e:
            print(f"  Error searching for {hotel_name}: {e}")
            return {
                'hotel_name': hotel_name,
                'price': None,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
    
    def _extract_price_from_hotel_page(self, hotel_name: str, found_name: str, url: str) -> Optional[Dict]:
        """Extract price from a hotel page"""
        try:
            print(f"    Extracting price from hotel page: {url}")
            price = None
            found_name = ""
            
            # Try to get hotel name from page
            try:
                name_selectors = ["h2.pc-header__title", ".hp__hotel-name", "h1", "[data-testid='hotel-name']"]
                for selector in name_selectors:
                    try:
                        name_elem = self.driver.find_element(By.CSS_SELECTOR, selector)
                        found_name = name_elem.text.strip()
                        if found_name:
                            break
                    except:
                        continue
            except:
                pass
            
            # Try to find price
            price_selectors = [
                "span[data-testid='price-and-discounted-price']",
                ".bui-price-display__value",
                ".prco-valign-middle-helper",
                ".bui-price-display",
                "[data-testid='price']",
                ".hprt-price-price",
                ".bui-price-display__value"
            ]
            
            for selector in price_selectors:
                try:
                    price_elem = WebDriverWait(self.driver, 8).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    price_text = price_elem.text.strip()
                    price = self._extract_price(price_text)
                    if price and price > 50:
                        break
                except:
                    continue
            
            if price:
                return {
                    'hotel_name': hotel_name,
                    'found_name': found_name or hotel_name,
                    'price': price,
                    'currency': 'QAR',
                    'url': url,
                    'timestamp': datetime.now().isoformat()
                }
        except Exception as e:
            print(f"    Error extracting from hotel page: {e}")
        return None
    
    def _extract_price_from_search_results(self, hotel_name: str, search_query: str) -> Optional[Dict]:
        """Extract price from search results page"""
        try:
            # Check if we're already on a hotel page (booking.com sometimes redirects directly)
            current_url = self.driver.current_url
            if '/hotel/' in current_url and '/searchresults' not in current_url:
                # We're on a hotel page directly - extract price from here
                try:
                    print(f"    Found hotel page directly: {current_url}")
                    price = None
                    found_name = ""
                    
                    # Try to get hotel name from page
                    try:
                        name_selectors = ["h2.pc-header__title", ".hp__hotel-name", "h1", "[data-testid='hotel-name']"]
                        for selector in name_selectors:
                            try:
                                name_elem = self.driver.find_element(By.CSS_SELECTOR, selector)
                                found_name = name_elem.text.strip()
                                if found_name:
                                    break
                            except:
                                continue
                    except:
                        pass
                    
                    # Try to find price
                    price_selectors = [
                        "span[data-testid='price-and-discounted-price']",
                        ".bui-price-display__value",
                        ".prco-valign-middle-helper",
                        ".bui-price-display",
                        "[data-testid='price']",
                        ".hprt-price-price",
                        ".bui-price-display__value"
                    ]
                    
                    for selector in price_selectors:
                        try:
                            price_elem = WebDriverWait(self.driver, 8).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                            )
                            price_text = price_elem.text.strip()
                            price = self._extract_price(price_text)
                            if price and price > 50:
                                break
                        except:
                            continue
                    
                    if price:
                        return {
                            'hotel_name': hotel_name,
                            'found_name': found_name or hotel_name,
                            'price': price,
                            'currency': 'QAR',
                            'url': current_url,
                            'timestamp': datetime.now().isoformat()
                        }
                except Exception as e:
                    print(f"    Error extracting from hotel page: {e}")
            
            # Try to find hotel in results
            try:
                # Multiple selectors for hotel links
                hotel_selectors = [
                    "a[data-testid='title-link']",
                    "a[data-testid='property-card-link']",
                    ".sr_item",
                    "[data-testid='property-card']"
                ]
                
                hotel_links = []
                for selector in hotel_selectors:
                    try:
                        links = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        if links:
                            hotel_links = links
                            break
                    except:
                        continue
                
                if not hotel_links:
                    # Try alternative: look for any hotel listing
                    hotel_links = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='/hotel/']")
                
                best_match = None
                best_score = 0
                
                # Extract key identifying words from hotel name (remove common words)
                common_words = {'فندق', 'hotel', 'الدوحة', 'doha', 'فنادق', 'اجنحة', 'apartments', 'residence', 'residences', 'ريزيدنس', 'ريزيدنسز', 'ومنتجع', 'resort'}
                hotel_words = set(word.lower() for word in hotel_name.split() if word.lower() not in common_words and len(word) > 2)
                
                # Also create a simplified search string (key words only)
                hotel_key_search = ' '.join(sorted(hotel_words))
                hotel_lower = hotel_name.lower()
                
                for link in hotel_links[:15]:  # Check first 15 results
                    try:
                        # Re-find the link to avoid stale element
                        link_text = ""
                        try:
                            link_text = link.text.strip()
                        except:
                            # If stale, try to get from href
                            try:
                                href = link.get_attribute('href')
                                if href:
                                    # Try to find the element again by href
                                    link = self.driver.find_element(By.CSS_SELECTOR, f"a[href='{href}']")
                                    link_text = link.text.strip()
                            except:
                                continue
                        
                        if not link_text:
                            continue
                        
                        link_lower = link_text.lower()
                        
                        # Multiple matching strategies
                        score = 0
                        matches = 0
                        
                        # Strategy 1: Key word matching
                        if hotel_words:
                            matches = sum(1 for word in hotel_words if word in link_lower)
                            score = matches / len(hotel_words) if hotel_words else 0
                        
                        # Strategy 2: Check for significant substring match (for longer names)
                        if len(hotel_lower) > 8:
                            # Try different length substrings
                            for sublen in [12, 10, 8]:
                                hotel_sub = hotel_lower[:sublen]
                                if hotel_sub in link_lower:
                                    score = max(score, 0.5 + (sublen / 20))
                                    break
                        
                        # Strategy 3: Check if key unique words appear
                        unique_keywords = [w for w in hotel_words if len(w) > 4]  # Longer, more unique words
                        if unique_keywords:
                            unique_matches = sum(1 for word in unique_keywords if word in link_lower)
                            if unique_matches > 0:
                                score = max(score, 0.4 + (unique_matches * 0.2))
                        
                        # Strategy 4: For English hotel names, check brand names
                        if 'four' in hotel_lower or 'seasons' in hotel_lower:
                            if 'four' in link_lower and 'season' in link_lower:
                                score = max(score, 0.7)
                        if 'marriott' in hotel_lower and 'marriott' in link_lower:
                            score = max(score, 0.7)
                        if 'sheraton' in hotel_lower and 'sheraton' in link_lower:
                            score = max(score, 0.7)
                        if 'intercontinental' in hotel_lower or 'inter' in hotel_lower:
                            if 'intercontinental' in link_lower or 'inter' in link_lower:
                                score = max(score, 0.7)
                        if 'hyatt' in hotel_lower or 'حياة' in hotel_lower:
                            if 'hyatt' in link_lower or 'grand hyatt' in link_lower:
                                score = max(score, 0.7)
                        
                        # Lower threshold - accept if we have any reasonable match
                        if score > best_score and (score >= 0.3 or matches >= 1):
                            best_score = score
                            best_match = link
                            if score >= 0.6:  # Very good match, use it
                                break
                    except Exception as e:
                        continue
                
                if best_match and best_score >= 0.3:  # Lower threshold for matching
                    try:
                        # Store the found name before accessing attributes (avoid stale element)
                        found_name = ""
                        hotel_url = ""
                        
                        try:
                            found_name = best_match.text.strip()[:100]
                            hotel_url = best_match.get_attribute('href')
                        except:
                            # If stale, re-find by index or other method
                            # Get URL from the search results page directly
                            try:
                                # Find all links again and get the one at the same position
                                all_links = self.driver.find_elements(By.CSS_SELECTOR, "a[data-testid='title-link']")
                                if all_links and len(all_links) > hotel_links.index(best_match) if best_match in hotel_links else 0:
                                    idx = hotel_links.index(best_match) if best_match in hotel_links else 0
                                    best_match = all_links[idx]
                                    found_name = best_match.text.strip()[:100]
                                    hotel_url = best_match.get_attribute('href')
                            except:
                                pass
                        
                        if not hotel_url or not hotel_url.startswith('http'):
                            # Try to find URL in parent
                            try:
                                hotel_url = best_match.find_element(By.XPATH, "./ancestor::a").get_attribute('href')
                            except:
                                pass
                        
                        if hotel_url:
                            # Try to get price from search results first
                            price = None
                            try:
                                # Look for price in the same card/container - use page source instead of element
                                page_source = self.driver.page_source
                                
                                # Try to find price near the hotel name in the HTML
                                price_selectors = [
                                    "span[data-testid='price-and-discounted-price']",
                                    ".bui-price-display__value",
                                    ".prco-valign-middle-helper",
                                    ".bui-price-display",
                                    "[data-testid='price']",
                                    ".sr_price"
                                ]
                                
                                # Find price element by searching near the hotel link
                                for selector in price_selectors:
                                    try:
                                        # Find all price elements
                                        price_elems = self.driver.find_elements(By.CSS_SELECTOR, selector)
                                        # Get the one closest to our hotel link
                                        if price_elems:
                                            # Use the first price found (they're usually in order)
                                            price_elem = price_elems[0]
                                            price_text = price_elem.text.strip()
                                            price = self._extract_price(price_text)
                                            if price and price > 50:  # Reasonable price check
                                                break
                                    except:
                                        continue
                            except:
                                pass
                            
                            # If price not found in search results, go to hotel page
                            if not price:
                                try:
                                    self.driver.get(hotel_url)
                                    time.sleep(5)  # Wait longer
                                    
                                    price_selectors = [
                                        "span[data-testid='price-and-discounted-price']",
                                        ".bui-price-display__value",
                                        ".prco-valign-middle-helper",
                                        ".bui-price-display",
                                        "[data-testid='price']",
                                        ".hprt-price-price"
                                    ]
                                    
                                    for selector in price_selectors:
                                        try:
                                            price_elem = WebDriverWait(self.driver, 8).until(
                                                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                                            )
                                            price_text = price_elem.text.strip()
                                            price = self._extract_price(price_text)
                                            if price and price > 50:  # Reasonable price check
                                                break
                                        except:
                                            continue
                                except Exception as e:
                                    pass
                            
                            if price and found_name:
                                # Return if we have a price and reasonable match
                                return {
                                    'hotel_name': hotel_name,
                                    'found_name': found_name,
                                    'price': price,
                                    'currency': 'QAR',
                                    'url': hotel_url,
                                    'timestamp': datetime.now().isoformat()
                                }
                    except Exception as e:
                        print(f"    Error extracting details: {e}")
                
                # Hotel not found or price unavailable
                return {
                    'hotel_name': hotel_name,
                    'found_name': None,
                    'price': None,
                    'currency': None,
                    'url': None,
                    'timestamp': datetime.now().isoformat(),
                    'error': 'Hotel not found or price unavailable'
                }
                
            except Exception as e:
                print(f"  Error finding hotel {hotel_name}: {e}")
                return {
                    'hotel_name': hotel_name,
                    'price': None,
                    'error': str(e),
                    'timestamp': datetime.now().isoformat()
                }
                
        except Exception as e:
            print(f"  Error searching for {hotel_name}: {e}")
            return {
                'hotel_name': hotel_name,
                'price': None,
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
    
    def _extract_price(self, price_text: str) -> Optional[float]:
        """Extract numeric price from text"""
        try:
            # Remove currency symbols and extract numbers
            import re
            # Remove common currency symbols and text
            cleaned = re.sub(r'[^\d.,]', '', price_text)
            cleaned = cleaned.replace(',', '')
            price = float(cleaned)
            return price
        except:
            return None
    
    def scrape_all_hotels(self) -> Dict:
        """Scrape prices for all hotels"""
        print("\n" + "="*60)
        print("HOTEL PRICE SCRAPER - BOOKING.COM")
        print("="*60)
        print(f"Scraping {len(self.hotels)} hotels...")
        print("="*60 + "\n")
        
        self._setup_driver(headless=self.headless)
        
        results = {
            'timestamp': datetime.now().isoformat(),
            'source': 'booking.com',
            'location': 'Doha, Qatar',
            'hotels': []
        }
        
        for i, hotel in enumerate(self.hotels, 1):
            print(f"[{i}/{len(self.hotels)}] Searching for: {hotel}")
            hotel_data = self._search_hotel(hotel)
            
            if hotel_data:
                results['hotels'].append(hotel_data)
                if hotel_data.get('price'):
                    print(f"  ✓ Found: {hotel_data.get('found_name', hotel)} - Price: {hotel_data.get('price')} {hotel_data.get('currency', '')}")
                else:
                    print(f"  ✗ Not found or price unavailable")
            else:
                results['hotels'].append({
                    'hotel_name': hotel,
                    'price': None,
                    'error': 'Search failed',
                    'timestamp': datetime.now().isoformat()
                })
                print(f"  ✗ Search failed")
            
            # Delay between requests to avoid being blocked
            time.sleep(3)
        
        self._close_driver()
        
        # Print summary
        found_count = sum(1 for h in results['hotels'] if h.get('price'))
        print("\n" + "="*60)
        print(f"Scraping completed: {found_count}/{len(self.hotels)} hotels found")
        print("="*60)
        
        return results
    
    def export_to_excel(self, results: Dict, filename: str = 'hotel_prices.xlsx'):
        """Export hotel prices to Excel file in RTL table format"""
        if 'error' in results:
            print(f"Error: Cannot export - {results['error']}")
            return False
        
        # Create or load workbook
        hotel_col = 1  # Column A contains hotel names (appears rightmost in RTL view)
        week_label_row = 1  # Row 1: Week label
        date_row = 2  # Row 2: Date header
        
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            ws.sheet_view.rightToLeft = True
            max_col = ws.max_column
            if max_col < 1:
                max_col = 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Hotel Prices"
            ws.sheet_view.rightToLeft = True
            
            # Initialize hotel names in column A
            for i, hotel_data in enumerate(results.get('hotels', []), 3):
                hotel_name = hotel_data.get('hotel_name', '')
                label_cell = ws.cell(row=i, column=hotel_col)
                label_cell.value = hotel_name
                label_cell.font = Font(bold=True)
                label_cell.alignment = Alignment(horizontal='right', vertical='center')
                label_cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
            
            # Add "الفندق" header in row 2, column A
            header_label_cell = ws.cell(row=date_row, column=hotel_col)
            header_label_cell.value = 'الفندق'  # Hotel
            header_label_cell.font = Font(bold=True)
            header_label_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_label_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            max_col = 1
        
        # Get current week
        today = datetime.now()
        week_start = today - timedelta(days=today.weekday())
        week_label = f"أسبوع {week_start.strftime('%Y-%m-%d')}"
        date_text = today.strftime('%Y-%m-%d')
        
        # Check if this week's column already exists
        date_col = None
        for col in range(2, max_col + 2):
            cell_value = ws.cell(row=date_row, column=col).value
            if cell_value and date_text in str(cell_value):
                date_col = col
                break
        
        if date_col is None:
            last_date_col = max_col
            if last_date_col < 2:
                last_date_col = 1
            date_col = last_date_col + 1
        
        # Set row 1: Week label
        week_cell = ws.cell(row=week_label_row, column=date_col)
        week_cell.value = week_label
        week_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        week_cell.font = Font(bold=True)
        week_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set row 2: Date
        date_cell = ws.cell(row=date_row, column=date_col)
        date_cell.value = date_text
        date_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ensure "الفندق" header is in column A, row 2
        header_label_cell = ws.cell(row=date_row, column=hotel_col)
        if header_label_cell.value is None or header_label_cell.value != 'الفندق':
            header_label_cell.value = 'الفندق'
        header_label_cell.font = Font(bold=True)
        header_label_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write hotel prices
        for i, hotel_data in enumerate(results.get('hotels', []), 3):
            hotel_name = hotel_data.get('hotel_name', '')
            
            # Ensure hotel name is in column A
            label_cell = ws.cell(row=i, column=hotel_col)
            if label_cell.value is None:
                label_cell.value = hotel_name
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            label_cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
            
            # Write price
            price = hotel_data.get('price')
            price_cell = ws.cell(row=i, column=date_col)
            if price:
                price_cell.value = price
                price_cell.number_format = '0.00'
            price_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ensure we have enough rows
        num_hotels = len(results.get('hotels', []))
        for row_num in range(3, num_hotels + 3):
            if ws.max_row < row_num:
                ws.append([None] * ws.max_column)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30  # Hotel name column
        for col in range(2, date_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        ws.sheet_view.rightToLeft = True
        wb.save(filename)
        print(f"\nPrices exported to {filename} (RTL layout)")
        return True


def main():
    """Main function"""
    import sys
    
    # Set headless=False to see the browser, True to run in background
    scraper = HotelPriceScraper(headless=False)
    
    # If hotel name provided as argument, test with just that hotel
    if len(sys.argv) > 1:
        test_hotel = sys.argv[1]
        print(f"Testing with single hotel: {test_hotel}\n")
        scraper.hotels = [test_hotel]
    
    results = scraper.scrape_all_hotels()
    
    # Save to JSON
    json_file = 'hotel_prices.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nPrices saved to {json_file}")
    
    # Export to Excel
    excel_file = 'hotel_prices.xlsx'
    scraper.export_to_excel(results, excel_file)


if __name__ == "__main__":
    main()
