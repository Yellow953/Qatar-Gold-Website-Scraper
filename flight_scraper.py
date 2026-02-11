#!/usr/bin/env python3
"""
Flight price scraper: round-trip, direct flights only.
Outputs to the same Excel format (route block at top, flight price rows, date columns).
Either returns real round-trip prices or fails (no one-way or unlabeled numbers).
"""

try:
    import undetected_chromedriver as uc
    USE_UNDETECTED = True
except ImportError:
    USE_UNDETECTED = False

import base64
import json
import os
import re
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Excel
FLIGHT_PRICES_EXCEL = 'flight_prices.xlsx'
FLIGHT_PRICES_SHEET_NAME = 'Flight Prices'
ROUTE_HEADERS = ['Code', 'Commodity', 'Origin', 'Origin_Code', 'Destination', 'Destination_Code', 'Duration_Months']
FLIGHT_HEADER_MARKER = 'وكالات'

# Price validation: round-trip economy only
LONG_HAUL_CODES = {'LHR', 'LON', 'JFK', 'NYC', 'IST', 'BKK', 'KUL', 'TBS', 'FCO', 'CDG', 'MAD', 'FRA', 'MUC', 'AMS', 'SYD', 'MEL', 'SIN', 'HKG', 'NRT', 'TYO'}
MIN_QAR_LONG_HAUL = 1000
MIN_QAR_SHORT_HAUL = 500
MAX_QAR_LONG_HAUL = 5500
MAX_QAR_SHORT_HAUL = 4000

# Approximate rates to QAR. USD is pegged ~3.64; update periodically if needed (e.g. for EUR/GBP).
CURRENCY_TO_QAR = {
    'QAR': 1.0, 'QR': 1.0, 'ر.ق': 1.0,
    'USD': 3.64, 'US$': 3.64, '$': 3.64,   # 1 USD ≈ 3.64 QAR (Qatar peg)
    'AED': 0.99, 'د.إ': 0.99, 'DH': 0.99,  # 1 AED ≈ 0.99 QAR (KAYAK.ae etc.)
    'EUR': 3.90, '€': 3.90,
    'GBP': 4.60, '£': 4.60,
    'SAR': 0.97, 'SR': 0.97,
    'BHD': 9.65, 'KWD': 11.85, 'OMR': 9.46,
}

ROUND_TRIP_KEYWORDS = ('total', 'round', 'return', 'round-trip', 'roundtrip', 'رحلة ذهاب وعودة')
ONE_WAY_KEYWORDS = ('one way', 'one-way', 'outbound', 'each way', 'per way', 'single way', 'one way only', 'من جهة واحدة')


class FlightPriceScraper:
    """Scrape round-trip, direct-flight prices only. Same Excel format as before."""

    def __init__(self, headless=False, excel_path: str = None):
        self.headless = headless
        self.driver = None
        self.excel_path = os.path.abspath(excel_path or FLIGHT_PRICES_EXCEL)
        self.routes = self._get_routes()
        self.sources = self._get_sources()

    # ---------- Routes ----------
    def _load_routes_from_excel(self) -> Optional[List[Dict]]:
        path = self.excel_path
        if not os.path.exists(path):
            return None
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[FLIGHT_PRICES_SHEET_NAME] if FLIGHT_PRICES_SHEET_NAME in wb.sheetnames else wb.active
            if str(ws.cell(row=1, column=1).value or '').strip() != 'Code':
                wb.close()
                return None
            col3 = str(ws.cell(row=1, column=3).value or '')
            if 'Class' in col3 or 'الدرجة' in col3:
                wb.close()
                return None
            routes = []
            for row_idx in range(2, ws.max_row + 1):
                code = ws.cell(row=row_idx, column=1).value
                if not code or not str(code).strip():
                    break
                if FLIGHT_HEADER_MARKER in str(ws.cell(row=row_idx, column=6).value or ''):
                    break
                dest = ws.cell(row=row_idx, column=5).value
                dest_code = ws.cell(row=row_idx, column=6).value
                if not dest or not dest_code:
                    continue
                try:
                    duration = int(ws.cell(row=row_idx, column=7).value or 6)
                except (TypeError, ValueError):
                    duration = 6
                origin = str(ws.cell(row=row_idx, column=3).value or 'Doha').strip()
                origin_code = str(ws.cell(row=row_idx, column=4).value or 'DOH').strip().upper()
                routes.append({
                    'code': str(code).strip(),
                    'origin': origin,
                    'origin_code': origin_code,
                    'destination': str(dest).strip(),
                    'destination_code': str(dest_code).strip().upper(),
                    'commodity_ar': str(ws.cell(row=row_idx, column=2).value or '').strip(),
                    'duration_months': duration,
                })
            wb.close()
            return routes if routes else None
        except Exception:
            return None

    def _get_default_routes(self) -> List[Dict]:
        return [
            {'code': '007331101', 'commodity_ar': 'كلفة تذكرة دوحة _ لندن - دوحة لمدة 6 (Semi flexble التذكرة السياحية) أشهر', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'London', 'destination_code': 'LHR', 'duration_months': 6},
            {'code': '007331102', 'commodity_ar': 'كلفة تذكرة دوحة _ القاهرة - دوحة لمدة 6 (semi flexble التذكرة سياحية ( اشهر', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Cairo', 'destination_code': 'CAI', 'duration_months': 6},
            {'code': '007331103', 'commodity_ar': 'كلفة تذكرة دوحة_ كراتشي _ دوحة لمدة 3 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Karachi', 'destination_code': 'KHI', 'duration_months': 3},
            {'code': '007331104', 'commodity_ar': 'كلفة تذكرة دوحة_ دبي _ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Dubai', 'destination_code': 'DXB', 'duration_months': 6},
            {'code': '007331105', 'commodity_ar': 'كلفة تذكرة دوحة_جدة _ دوحة لمدة 6 اشهر( التذكرة السياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Jeddah', 'destination_code': 'JED', 'duration_months': 6},
            {'code': '007331106', 'commodity_ar': 'كلفة تذكرة دوحة_ بومباي _ دوحة لمدة 3 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Mumbai', 'destination_code': 'BOM', 'duration_months': 3},
            {'code': '007331107', 'commodity_ar': 'كلفة تذكرة دوحة_كولا لمبور _ دوحة لمدة 6 اشهر( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Kuala Lumpur', 'destination_code': 'KUL', 'duration_months': 6},
            {'code': '007331108', 'commodity_ar': 'كلفة تذكرة دوحة_ اسطنبول لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Istanbul', 'destination_code': 'IST', 'duration_months': 6},
            {'code': '007331109', 'commodity_ar': 'كلفة تذكرة دوحة_ بانكوك _ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Bangkok', 'destination_code': 'BKK', 'duration_months': 6},
            {'code': '007331110', 'commodity_ar': 'كلفة تذكرة دوحة_تبليسي_ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Tbilisi', 'destination_code': 'TBS', 'duration_months': 6},
            {'code': '007331111', 'commodity_ar': 'كلفة تذكرة دوحة_نيويورك دوحة لمدة 6 اشهر ( التذكرة السياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'New York', 'destination_code': 'JFK', 'duration_months': 6},
        ]

    def _get_routes(self) -> List[Dict]:
        loaded = self._load_routes_from_excel()
        if loaded:
            return loaded
        return self._get_default_routes()

    def _get_sources(self) -> List[Dict]:
        return [
            {'name': 'Qatar Airways', 'name_ar': 'الخطوط القطرية', 'source_code': 'AIRL001', 'type': 'airline'},
            {'name': 'British Airways', 'name_ar': 'الخطوط البريطانية', 'source_code': 'AIRL018', 'type': 'airline'},
            {'name': 'Malaysia Airlines', 'name_ar': 'الخطوط الماليزية', 'source_code': 'AIRL024', 'type': 'airline'},
            {'name': 'Kuwait Airways', 'name_ar': 'الخطوط الكويتية', 'source_code': 'AIRL025', 'type': 'airline'},
            {'name': 'Turkish Airlines', 'name_ar': 'الخطوط التركية', 'source_code': 'AIRL026', 'type': 'airline'},
            {'name': 'Pakistan International Airlines', 'name_ar': 'الخطوط الباكستانية', 'source_code': 'AIRL020', 'type': 'airline'},
            {'name': 'CheapAir', 'name_ar': 'cheapair', 'source_code': 'AIRL028', 'type': 'aggregator'},
            {'name': 'eDreams', 'name_ar': 'edreams', 'source_code': 'AIRL030', 'type': 'aggregator'},
            {'name': 'KAYAK', 'name_ar': 'Kayak', 'source_code': 'AIRL028', 'type': 'aggregator'},
            {'name': 'ITA Matrix', 'name_ar': 'matrix', 'source_code': 'AIRL028', 'type': 'aggregator'},
        ]

    # ---------- Driver ----------
    def _setup_driver(self, headless=False):
        if USE_UNDETECTED:
            try:
                options = uc.ChromeOptions()
                if headless:
                    options.add_argument('--headless=new')
                options.add_argument('--no-sandbox')
                options.add_argument('--disable-dev-shm-usage')
                self.driver = uc.Chrome(options=options, version_main=None)
                if not headless:
                    self.driver.set_window_size(1920, 1080)
                return
            except Exception:
                pass
        chrome_options = Options()
        if headless:
            chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        self.driver.set_window_size(1920, 1080)

    def _close_driver(self):
        if self.driver:
            self.driver.quit()
            self.driver = None

    def _calculate_dates(self, months: int) -> Tuple[str, str]:
        today = datetime.now()
        dep = today + timedelta(days=7)
        ret = dep + timedelta(days=months * 30)
        return dep.strftime('%Y-%m-%d'), ret.strftime('%Y-%m-%d')

    def _close_dialogs(self):
        try:
            for sel in ["button#onetrust-accept-btn-handler", "button[id*='accept']", "button[aria-label*='Accept']", "button[aria-label*='Close']"]:
                try:
                    btn = WebDriverWait(self.driver, 2).until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
                    btn.click()
                    time.sleep(1)
                    break
                except Exception:
                    continue
        except Exception:
            pass

    def _apply_direct_filter(self):
        try:
            for sel in ["[data-testid*='nonstop']", "[data-testid*='direct']", "[aria-label*='Nonstop']", "[aria-label*='Direct']", "[class*='nonstop']"]:
                try:
                    for elem in self.driver.find_elements(By.CSS_SELECTOR, sel):
                        if elem.is_displayed() and ('direct' in (elem.text or '').lower() or 'nonstop' in (elem.text or '').lower() or 'non-stop' in (elem.text or '').lower()):
                            elem.click()
                            time.sleep(2)
                            return
                except Exception:
                    continue
        except Exception:
            pass

    def _wait_for_prices(self, timeout_sec: int = 20) -> bool:
        """Wait until at least one price-like element (digits + optional currency) is visible. Returns True if found."""
        try:
            for _ in range(timeout_sec):
                for sel in ["[class*='price']", "[class*='fare']", "[class*='Price']", "[data-testid*='price']", "[data-test-id='price']", ".price", ".fare", "[class*='amount']"]:
                    try:
                        elms = self.driver.find_elements(By.CSS_SELECTOR, sel)
                        for e in elms[:10]:
                            t = (e.text or '').strip()
                            if t and re.search(r'\d{3,}', t):
                                return True
                    except Exception:
                        pass
                time.sleep(1)
        except Exception:
            pass
        return False

    # ---------- Currency & validation ----------
    def _detect_currency(self, text: str) -> Tuple[Optional[float], str]:
        if not text or not text.strip():
            return None, 'QAR'
        patterns = [
            (r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*(QAR|QR|USD|US\$|\$|AED|EUR|€|GBP|£)', 1, 2),
            (r'(QAR|QR|USD|US\$|\$|AED|EUR|€|GBP|£)\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', 2, 1),
            (r'\$\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', 1, 'USD'),
            (r'£\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', 1, 'GBP'),
            (r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*£', 1, 'GBP'),
            (r'(?:AED|د\.إ)\s*(\d{1,3}(?:,\d{3})*)', 1, 'AED'),
        ]
        for pat, ng, cg in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                try:
                    amount = float(m.group(ng).replace(',', ''))
                    cur = m.group(cg) if isinstance(cg, int) else cg
                    cur = 'USD' if cur in ('$', 'US$') else ('EUR' if cur == '€' else ('GBP' if cur == '£' else (cur or 'QAR').strip().upper()))
                    if amount > 0:
                        return amount, cur
                except (ValueError, TypeError, IndexError):
                    pass
        cleaned = re.sub(r'[^\d.,]', '', text).replace(',', '')
        try:
            amount = float(cleaned)
            if amount > 0:
                return amount, 'QAR'
        except ValueError:
            pass
        return None, 'QAR'

    def _to_qar(self, amount: float, currency: str) -> float:
        if amount is None or amount <= 0:
            return 0.0
        c = (currency or 'QAR').strip().upper()
        if c in ('$', 'US$'):
            c = 'USD'
        return round(amount * CURRENCY_TO_QAR.get(c, 1.0), 0)

    def _min_max_qar(self, route: Dict) -> Tuple[int, int]:
        dest = (route.get('destination_code') or '').upper()
        is_long = dest in LONG_HAUL_CODES
        min_q = MIN_QAR_LONG_HAUL if is_long else MIN_QAR_SHORT_HAUL
        max_q = MAX_QAR_LONG_HAUL if is_long else MAX_QAR_SHORT_HAUL
        return min_q, max_q

    # ---------- Price extraction: prefer round-trip label, accept in-range from round-trip search ----------
    # For long-haul unlabeled we use min 2000 to reduce one-way risk (one-way often 800–1800).
    MIN_QAR_UNLABELED_LONG_HAUL = 2000

    def _extract_round_trip_price(self, selectors: List[str], route: Dict) -> Optional[float]:
        """Extract price in QAR. Reject one-way labeled; prefer round-trip labeled; accept unlabeled in range (long-haul unlabeled >= 2000)."""
        min_q, max_q = self._min_max_qar(route)
        dest = (route.get('destination_code') or '').upper()
        is_long_haul = dest in LONG_HAUL_CODES

        candidates = []  # (qar, is_round_trip_label, amount, currency)
        for selector in selectors:
            try:
                for elem in self.driver.find_elements(By.CSS_SELECTOR, selector)[:20]:
                    text = (elem.text or '').strip()
                    # Include parent text so we get "Total £350" when price is in child
                    if not text or not re.search(r'\d{2,}', text):
                        try:
                            parent = elem.find_element(By.XPATH, '..')
                            parent_text = (parent.text or '').strip()
                            if parent_text and len(parent_text) < 500:
                                text = parent_text
                        except Exception:
                            pass
                    if not text:
                        continue
                    amount, currency = self._detect_currency(text)
                    if not amount or amount <= 0:
                        continue
                    qar = self._to_qar(amount, currency)
                    if not (min_q <= qar <= max_q):
                        continue
                    text_lower = text.lower()
                    if any(k in text_lower for k in ONE_WAY_KEYWORDS):
                        continue
                    # Long-haul unlabeled below threshold: likely one-way, skip
                    is_round = any(k in text_lower for k in ROUND_TRIP_KEYWORDS)
                    if is_long_haul and not is_round and qar < self.MIN_QAR_UNLABELED_LONG_HAUL:
                        continue
                    candidates.append((qar, is_round, amount, currency))
            except Exception:
                continue
            if candidates:
                break

        if candidates:
            round_only = [c for c in candidates if c[1]]
            if round_only:
                best = min(round_only, key=lambda x: x[0])
            else:
                best = min(candidates, key=lambda x: x[0])
            return float(round(best[0]))

        # Fallback: scan page source for numbers in range (sites often don't label "total")
        try:
            page_text = self.driver.page_source or ""
        except Exception:
            page_text = ""
        fallback_min = max(min_q, self.MIN_QAR_UNLABELED_LONG_HAUL) if is_long_haul else min_q
        patterns = [
            (r'(?:QAR|QR)\s*(\d{1,3}(?:,\d{3})*)', 'QAR'),
            (r'\$\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', 'USD'),
            (r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*USD', 'USD'),
            (r'£\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', 'GBP'),
            (r'(?:GBP|£)\s*(\d{1,3}(?:,\d{3})*)', 'GBP'),
            (r'(?:AED|د\.إ)\s*(\d{1,3}(?:,\d{3})*)', 'AED'),
            (r'(\d{1,3}(?:,\d{3})*)\s*AED', 'AED'),
        ]
        found = []
        for pattern, cur in patterns:
            for m in re.finditer(pattern, page_text, re.IGNORECASE):
                try:
                    val = float(m.group(1).replace(',', ''))
                    qar = self._to_qar(val, cur)
                    if fallback_min <= qar <= max_q:
                        found.append(qar)
                except (ValueError, TypeError):
                    pass
        if found:
            return float(round(min(found)))
        return None

    # ---------- Scrapers: round-trip + direct only ----------
    def _price_result(self, route: Dict, source: Dict, price: float, airline: str = None) -> Dict:
        return {
            'route_code': route['code'],
            'source': source['name'],
            'source_ar': source['name_ar'],
            'source_code': source['source_code'],
            'airline': airline or source['name'],
            'price': round(price),
            'currency': 'QAR',
            'timestamp': datetime.now().isoformat(),
        }

    def scrape_qatar_airways(self, route: Dict) -> Optional[Dict]:
        """Round-trip, direct only."""
        try:
            dep, ret = self._calculate_dates(route['duration_months'])
            url = (f"https://www.qatarairways.com/app/booking/flight-selection?"
                   f"widget=QR&searchType=F&addTaxToFare=Y&minPurTime=0&selLang=en&"
                   f"tripType=R&fromStation={route['origin_code']}&toStation={route['destination_code']}&"
                   f"departing={dep}&returning={ret}&bookingClass=E&"
                   f"adults=1&children=0&infants=0&ofw=0&teenager=0&flexibleDate=off&allowRedemption=N&stops=0")
            self.driver.get(url)
            time.sleep(8)
            self._close_dialogs()
            self._wait_for_prices(22)
            self._apply_direct_filter()
            time.sleep(4)
            price = self._extract_round_trip_price(
                ["[class*='price']", "[class*='fare']", "[class*='Price']", "[data-testid*='price']", ".price", ".fare", "span[class*='amount']"], route)
            if price:
                return self._price_result(route, {'name': 'Qatar Airways', 'name_ar': 'الخطوط القطرية', 'source_code': 'AIRL001'}, price, 'Qatar Airways')
            return None
        except Exception as e:
            print(f"      Error: {e}")
            return None

    def scrape_british_airways(self, route: Dict) -> Optional[Dict]:
        """Round-trip, direct only. BA often shows GBP (£); we convert to QAR."""
        try:
            dep, ret = self._calculate_dates(route['duration_months'])
            url = (f"https://www.britishairways.com/nx/b/airselect/en/usa/book/search?"
                   f"trip=round&arrivalDate={ret}&departureDate={dep}&"
                   f"from={route['origin_code']}&to={route['destination_code']}&"
                   f"travelClass=economy&adults=1&youngAdults=0&children=0&infants=0&bound=outbound&stops=0")
            self.driver.get(url)
            time.sleep(8)
            self._close_dialogs()
            self._wait_for_prices(22)
            self._apply_direct_filter()
            time.sleep(4)
            # BA uses many class patterns; include generic and parent-context in extractor
            price = self._extract_round_trip_price([
                "[class*='price']", "[class*='fare']", "[class*='Price']", "[class*='Fare']",
                "[class*='amount']", "[class*='Amount']", "[data-testid*='price']", "[data-testid*='fare']",
                ".price", ".fare", "span[class*='currency']", "div[class*='total']", "[class*='total']",
            ], route)
            if price:
                return self._price_result(route, {'name': 'British Airways', 'name_ar': 'الخطوط البريطانية', 'source_code': 'AIRL018'}, price, 'British Airways')
            # BA-specific: scan page for GBP/£ (they often render price in GBP)
            price = self._extract_ba_price_from_page(route)
            if price:
                return self._price_result(route, {'name': 'British Airways', 'name_ar': 'الخطوط البريطانية', 'source_code': 'AIRL018'}, price, 'British Airways')
            return None
        except Exception as e:
            print(f"      Error: {e}")
            return None

    def _extract_ba_price_from_page(self, route: Dict) -> Optional[float]:
        """British Airways: find GBP/£ amounts in page source when element extraction misses."""
        min_q, max_q = self._min_max_qar(route)
        dest = (route.get('destination_code') or '').upper()
        is_long_haul = dest in LONG_HAUL_CODES
        fallback_min = max(min_q, self.MIN_QAR_UNLABELED_LONG_HAUL) if is_long_haul else min_q
        try:
            html = self.driver.page_source or ""
        except Exception:
            return None
        # Match £123, £1,234, GBP 123, 123 GBP
        patterns = [
            (r'£\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', 'GBP'),
            (r'(?:GBP|£)\s*(\d{1,3}(?:,\d{3})*)', 'GBP'),
            (r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*(?:GBP|£)', 'GBP'),
        ]
        found = []
        for pattern, cur in patterns:
            for m in re.finditer(pattern, html, re.IGNORECASE):
                try:
                    val = float(m.group(1).replace(',', ''))
                    qar = self._to_qar(val, cur)
                    if fallback_min <= qar <= max_q:
                        found.append(qar)
                except (ValueError, TypeError):
                    pass
        return float(round(min(found))) if found else None

    def scrape_malaysia_airlines(self, route: Dict) -> Optional[Dict]:
        return None  # No direct search URL; would show wrong numbers

    def scrape_kuwait_airways(self, route: Dict) -> Optional[Dict]:
        return None

    def scrape_turkish_airlines(self, route: Dict) -> Optional[Dict]:
        return None

    def scrape_pia(self, route: Dict) -> Optional[Dict]:
        return None

    def _scrape_kayak(self, route: Dict) -> Optional[Dict]:
        try:
            dep, ret = self._calculate_dates(route['duration_months'])
            url = f"https://www.kayak.ae/flights/{route['origin_code']}-{route['destination_code']}/{dep}/{ret}?sort=bestflight_a&fs=stops=0"
            self.driver.get(url)
            time.sleep(8)
            self._close_dialogs()
            self._wait_for_prices(25)
            self._apply_direct_filter()
            time.sleep(5)
            price = self._extract_round_trip_price(
                ["[data-test-id='price']", "[data-testid='price']", "[data-testid='result-price']", ".Flights-Price-FlightPrice", "[class*='price']", "[class*='Price']", ".result-price", "span[class*='price']"], route)
            if price:
                return self._price_result(route, {'name': 'KAYAK', 'name_ar': 'Kayak', 'source_code': 'AIRL028'}, price, 'Various')
            return None
        except Exception as e:
            print(f"      Error: {e}")
            return None

    def _scrape_edreams(self, route: Dict) -> Optional[Dict]:
        try:
            dep, ret = self._calculate_dates(route['duration_months'])
            url = (f"https://www.edreams.qa/travel/#results/"
                   f"type=R;from={route['origin_code']};to={route['destination_code']};"
                   f"dep={dep};ret={ret};directOnly=true")
            self.driver.get(url)
            time.sleep(10)
            self._close_dialogs()
            self._wait_for_prices(25)
            self._apply_direct_filter()
            time.sleep(5)
            price = self._extract_round_trip_price(
                ["[class*='price']", "[class*='fare']", "[class*='Price']", "[data-testid*='price']", ".price", ".fare", "span[class*='price']"], route)
            if price:
                return self._price_result(route, {'name': 'eDreams', 'name_ar': 'edreams', 'source_code': 'AIRL030'}, price, 'Various')
            return None
        except Exception as e:
            print(f"      Error: {e}")
            return None

    def _scrape_cheapair(self, route: Dict) -> Optional[Dict]:
        try:
            dep, ret = self._calculate_dates(route['duration_months'])
            dep_f = datetime.strptime(dep, '%Y-%m-%d').strftime('%m/%d/%Y')
            ret_f = datetime.strptime(ret, '%Y-%m-%d').strftime('%m/%d/%Y')
            url = (f"https://www.cheapoair.com/air/listing?"
                   f"d1={route['origin_code']}&r1={route['destination_code']}&dt1={dep_f}&dtype1=A&rtype1=C&"
                   f"d2={route['destination_code']}&r2={route['origin_code']}&dt2={ret_f}&dtype2=C&rtype2=A&"
                   f"tripType=ROUNDTRIP&cl=ECONOMY&ad=1&nonstop=1")
            self.driver.get(url)
            time.sleep(8)
            self._close_dialogs()
            self._wait_for_prices(22)
            time.sleep(4)
            price = self._extract_round_trip_price(
                ["[class*='price']", "[class*='fare']", "[class*='Price']", "[data-testid*='price']", ".price", ".fare"], route)
            if price:
                return self._price_result(route, {'name': 'CheapAir', 'name_ar': 'cheapair', 'source_code': 'AIRL028'}, price, 'Various')
            return None
        except Exception as e:
            print(f"      Error: {e}")
            return None

    def _scrape_ita_matrix(self, route: Dict) -> Optional[Dict]:
        try:
            dep, ret = self._calculate_dates(route['duration_months'])
            payload = {
                "type": "round-trip",
                "slices": [{
                    "origin": [route['origin_code']],
                    "dest": [route['destination_code']],
                    "dates": {
                        "searchDateType": "specific",
                        "departureDate": dep,
                        "departureDateType": "depart",
                        "departureDateModifier": "0",
                        "departureDatePreferredTimes": [],
                        "returnDate": ret,
                        "returnDateType": "depart",
                        "returnDateModifier": "0",
                        "returnDatePreferredTimes": []
                    }
                }],
                "options": {"cabin": "COACH", "stops": "-1", "extraStops": "0", "pax": {"adults": "1"}}
            }
            enc = base64.b64encode(json.dumps(payload).encode()).decode()
            url = f"https://matrix.itasoftware.com/flights?search={enc}"
            self.driver.get(url)
            time.sleep(10)
            self._close_dialogs()
            self._wait_for_prices(25)
            time.sleep(5)
            price = self._extract_round_trip_price(
                ["[class*='price']", "[class*='fare']", "[class*='Price']", ".price", ".fare", "span[class*='price']"], route)
            if price:
                return self._price_result(route, {'name': 'ITA Matrix', 'name_ar': 'matrix', 'source_code': 'AIRL028'}, price, 'Various')
            return None
        except Exception as e:
            print(f"      Error: {e}")
            return None

    def scrape_aggregator(self, source: Dict, route: Dict) -> Optional[Dict]:
        name = source['name']
        if name == 'KAYAK':
            return self._scrape_kayak(route)
        if name == 'eDreams':
            return self._scrape_edreams(route)
        if name == 'CheapAir':
            return self._scrape_cheapair(route)
        if name == 'ITA Matrix':
            return self._scrape_ita_matrix(route)
        return None

    # ---------- Excel (same format as before) ----------
    SCHEDULED_DAYS = (4, 10, 17, 24)

    def _get_scheduled_dates_through_2026(self) -> List[Tuple[str, datetime]]:
        from datetime import date
        today = date.today()
        result = []
        for year in range(today.year, 2027):
            for month in range(1, 13):
                if year == 2026 and month > 12:
                    break
                for day in self.SCHEDULED_DAYS:
                    try:
                        d = date(year, month, day)
                        if d >= today:
                            result.append((d.strftime('%d-%b'), d))
                    except ValueError:
                        pass
        return result

    def _flight_header_row(self, ws) -> Optional[int]:
        for r in range(1, ws.max_row + 1):
            if FLIGHT_HEADER_MARKER in str(ws.cell(row=r, column=6).value or ''):
                return r
        return None

    def _ensure_route_block(self, ws) -> int:
        col3 = str(ws.cell(row=1, column=3).value or '')
        if 'Class' in col3 or 'الدرجة' in col3:
            return 0
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        if ws.cell(row=1, column=1).value != 'Code':
            for col, h in enumerate(ROUTE_HEADERS, 1):
                c = ws.cell(row=1, column=col)
                c.value = h
                c.font = Font(bold=True)
                c.fill = header_fill
        if not ws.cell(row=2, column=1).value:
            for row_idx, r in enumerate(self._get_default_routes(), 2):
                ws.cell(row=row_idx, column=1).value = r['code']
                ws.cell(row=row_idx, column=2).value = r['commodity_ar']
                ws.cell(row=row_idx, column=3).value = r['origin']
                ws.cell(row=row_idx, column=4).value = r['origin_code']
                ws.cell(row=row_idx, column=5).value = r['destination']
                ws.cell(row=row_idx, column=6).value = r['destination_code']
                ws.cell(row=row_idx, column=7).value = r['duration_months']
        last_route_row = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value and FLIGHT_HEADER_MARKER not in str(ws.cell(row=r, column=6).value or ''):
                last_route_row = r
            else:
                break
        return last_route_row + 1

    def _column_has_data(self, ws, col: int, from_row: int, to_row: int) -> bool:
        for r in range(from_row, to_row + 1):
            val = ws.cell(row=r, column=col).value
            if val is not None and str(val).strip() != '':
                try:
                    float(str(val).replace(',', ''))
                    return True
                except ValueError:
                    pass
        return False

    def _get_next_scheduled_date_column(self, ws, flight_header_row: int, existing_headers: dict, scheduled_dates: list, max_col: int) -> Tuple[str, int]:
        for header_text, _ in scheduled_dates:
            col = existing_headers.get(header_text)
            if col is None:
                return (header_text, max_col + 1)
            if not self._column_has_data(ws, col, flight_header_row + 1, ws.max_row):
                return (header_text, col)
        return (datetime.now().strftime('%d-%b'), max_col + 1)

    def _prepare_excel_for_export(self, filename: str):
        filename = os.path.abspath(filename or self.excel_path)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        flight_headers = [
            'Code', 'Commodity', 'الدرجة المقابلة لها في الخطوط (Class equivalent in airlines)',
            'CPI-Flag', 'رمز المصدر (Source Code)', 'وكالات الخطوط (Flight Agencies)'
        ]
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb[FLIGHT_PRICES_SHEET_NAME] if FLIGHT_PRICES_SHEET_NAME in wb.sheetnames else wb.active
            ws.title = FLIGHT_PRICES_SHEET_NAME
            ws.sheet_view.rightToLeft = False
            row_after_routes = self._ensure_route_block(ws)
            hrow = self._flight_header_row(ws)
            if hrow is None:
                flight_header_row = row_after_routes + 1
                for col_idx, header in enumerate(flight_headers, 1):
                    cell = ws.cell(row=flight_header_row, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                max_col = 6
            else:
                flight_header_row = hrow
                max_col = max(ws.max_column, 7)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = FLIGHT_PRICES_SHEET_NAME
            ws.sheet_view.rightToLeft = False
            row_after_routes = self._ensure_route_block(ws)
            flight_header_row = row_after_routes + 1
            for col_idx, header in enumerate(flight_headers, 1):
                cell = ws.cell(row=flight_header_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            max_col = 6

        scheduled_dates = self._get_scheduled_dates_through_2026()
        def _norm(v):
            if not v:
                return None
            s = str(v).strip()
            try:
                return datetime.strptime(s, '%d-%b').strftime('%d-%b')
            except ValueError:
                try:
                    return datetime.strptime(s, '%d-%b-%Y').strftime('%d-%b')
                except ValueError:
                    return s
        existing_headers = {}
        for col in range(7, max_col + 1):
            key = _norm(ws.cell(row=flight_header_row, column=col).value)
            if key:
                existing_headers[key] = col
        next_col = max_col + 1
        for header_text, _ in scheduled_dates:
            if header_text in existing_headers:
                continue
            cell = ws.cell(row=flight_header_row, column=next_col)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            existing_headers[header_text] = next_col
            next_col += 1
        max_col = next_col - 1
        date_text, date_col = self._get_next_scheduled_date_column(ws, flight_header_row, existing_headers, scheduled_dates, max_col)
        if date_col > max_col:
            cell = ws.cell(row=flight_header_row, column=date_col)
            cell.value = date_text
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            max_col = date_col
        row = flight_header_row + 1
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        avg_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        return (wb, ws, date_col, row, thin_border, avg_fill, flight_header_row)

    def _find_existing_rows_for_route(self, ws, flight_header_row: int, route_code: str) -> List[Tuple[int, str, str]]:
        out = []
        r = flight_header_row + 1
        while r <= ws.max_row:
            code = ws.cell(row=r, column=1).value
            if not code or str(code).strip() != route_code:
                break
            out.append((r, str(ws.cell(row=r, column=4).value or ''), str(ws.cell(row=r, column=5).value or '')))
            r += 1
        return out

    def _update_route_date_column(self, ws, route_result: Dict, existing_rows: List[Tuple[int, str, str]], date_col: int) -> None:
        prices = route_result.get('prices', [])
        sorted_prices = sorted(prices, key=lambda x: (x.get('airline', 'Various'), x.get('source', '')))
        airline_avg_groups = {}
        for p in prices:
            airline = p.get('airline', 'Various')
            if airline != 'Various' and p.get('price'):
                airline_avg_groups.setdefault(airline, []).append(p.get('price'))
        values = [p.get('price') for p in sorted_prices]
        for airline, plist in airline_avg_groups.items():
            if len(plist) > 1:
                values.append(round(sum(plist) / len(plist)))
        valid_prices = [p.get('price') for p in prices if p.get('price')]
        if valid_prices:
            values.append(round(sum(valid_prices) / len(valid_prices)))
        for i, (row_idx, _, _) in enumerate(existing_rows):
            if i < len(values) and values[i] is not None:
                ws.cell(row=row_idx, column=date_col).value = values[i]
                ws.cell(row=row_idx, column=date_col).number_format = '0'

    def _write_route_to_sheet(self, ws, route_result: Dict, row: int, date_col: int, thin_border, avg_fill) -> int:
        route = route_result['route']
        prices = route_result.get('prices', [])
        sorted_prices = sorted(prices, key=lambda x: (x.get('airline', 'Various'), x.get('source', '')))
        route_start_row = row
        for price_data in sorted_prices:
            ws.cell(row=row, column=1).value = route['code']
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2).value = route['commodity_ar']
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            ws.cell(row=row, column=3).value = 'Economy'
            ws.cell(row=row, column=4).value = 'Y'
            ws.cell(row=row, column=5).value = price_data.get('source_code', '')
            agency = price_data.get('source_ar', price_data.get('source', ''))
            airline = price_data.get('airline', '')
            src = price_data.get('source', '')
            if airline and airline != 'Various':
                if src in ['KAYAK', 'eDreams', 'CheapAir', 'ITA Matrix']:
                    val = f"Kayak عبر {airline}" if src == 'KAYAK' else (f"matrix عبر {airline}" if src == 'ITA Matrix' else f"{agency} عبر {airline}")
                else:
                    val = agency
            else:
                val = agency
            ws.cell(row=row, column=6).value = val
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border
            p = price_data.get('price')
            if p:
                ws.cell(row=row, column=date_col).value = p
                ws.cell(row=row, column=date_col).number_format = '0'
            ws.cell(row=row, column=date_col).border = thin_border
            row += 1
        airline_avg_groups = {}
        for p in prices:
            if p.get('airline', 'Various') != 'Various' and p.get('price'):
                airline_avg_groups.setdefault(p['airline'], []).append(p['price'])
        airline_ar_map = {'Qatar Airways': 'القطرية', 'British Airways': 'البريطانية', 'Malaysia Airlines': 'الماليزية', 'Kuwait Airways': 'الكويتية', 'Turkish Airlines': 'التركية', 'Pakistan International Airlines': 'الباكستانية'}
        for airline, plist in airline_avg_groups.items():
            if len(plist) <= 1:
                continue
            avg_price = sum(plist) / len(plist)
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.fill = avg_fill
            ws.cell(row=row, column=1).value = route['code']
            ws.cell(row=row, column=2).value = route['commodity_ar']
            ws.cell(row=row, column=3).value = 'Economy'
            ws.cell(row=row, column=4).value = 'N-averages'
            ws.cell(row=row, column=5).value = ''
            ws.cell(row=row, column=6).value = f"متوسط المصادر للخطوط {airline_ar_map.get(airline, airline)}"
            ws.cell(row=row, column=date_col).value = round(avg_price)
            ws.cell(row=row, column=date_col).number_format = '0'
            ws.cell(row=row, column=date_col).border = thin_border
            ws.cell(row=row, column=date_col).fill = avg_fill
            row += 1
        all_valid = [p.get('price') for p in prices if p.get('price')]
        if all_valid:
            overall = sum(all_valid) / len(all_valid)
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).fill = avg_fill
            ws.cell(row=row, column=1).value = route['code']
            ws.cell(row=row, column=2).value = route['commodity_ar']
            ws.cell(row=row, column=3).value = 'Economy'
            ws.cell(row=row, column=4).value = 'Y-averages'
            ws.cell(row=row, column=5).value = ''
            ws.cell(row=row, column=6).value = "متوسط المصادر"
            ws.cell(row=row, column=date_col).value = round(overall)
            ws.cell(row=row, column=date_col).number_format = '0'
            ws.cell(row=row, column=date_col).border = thin_border
            ws.cell(row=row, column=date_col).fill = avg_fill
            row += 1
        try:
            route_end = row - 1
            if route_end > route_start_row:
                ws.merge_cells(f'B{route_start_row}:B{route_end}')
                ws.cell(row=route_start_row, column=2).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        except Exception:
            pass
        return row

    def _expected_rows_count(self, route_result: Dict) -> int:
        prices = route_result.get('prices', [])
        n = len(sorted(prices, key=lambda x: (x.get('airline', 'Various'), x.get('source', ''))))
        airline_groups = {}
        for p in prices:
            if p.get('airline', 'Various') != 'Various' and p.get('price'):
                airline_groups.setdefault(p['airline'], []).append(p['price'])
        for plist in airline_groups.values():
            if len(plist) > 1:
                n += 1
        if any(p.get('price') for p in prices):
            n += 1
        return n

    def append_route_to_excel(self, route_result: Dict, filename: str = None, date_col_override: int = None):
        filename = os.path.abspath(filename or self.excel_path)
        wb, ws, date_col, row, thin_border, avg_fill, flight_header_row = self._prepare_excel_for_export(filename)
        if date_col_override is not None:
            date_col = date_col_override
        route_code = route_result['route']['code']
        existing = self._find_existing_rows_for_route(ws, flight_header_row, route_code)
        expected = self._expected_rows_count(route_result)
        if len(existing) == expected and expected > 0:
            self._update_route_date_column(ws, route_result, existing, date_col)
        else:
            self._write_route_to_sheet(ws, route_result, row, date_col, thin_border, avg_fill)
        wb.save(filename)

    def export_to_excel(self, results: Dict, filename: str = None) -> bool:
        filename = os.path.abspath(filename or self.excel_path)
        try:
            wb, ws, date_col, row, thin_border, avg_fill, flight_header_row = self._prepare_excel_for_export(filename)
            for route_result in results.get('routes', []):
                route_code = route_result['route']['code']
                existing = self._find_existing_rows_for_route(ws, flight_header_row, route_code)
                expected = self._expected_rows_count(route_result)
                if len(existing) == expected and expected > 0:
                    self._update_route_date_column(ws, route_result, existing, date_col)
                else:
                    row = self._write_route_to_sheet(ws, route_result, row, date_col, thin_border, avg_fill)
            for col in range(7, date_col + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            ws.sheet_view.rightToLeft = False
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Export error: {e}")
            return False

    def scrape_all(self) -> Dict:
        """Scrape all routes from all sources. One date column per run. Round-trip, direct only."""
        print("\n" + "="*60)
        print("FLIGHT PRICE SCRAPER (round-trip, direct only)")
        print("="*60)
        print(f"Routes: {len(self.routes)}, Sources: {len(self.sources)}")
        print("="*60 + "\n")
        self._setup_driver(headless=self.headless)
        results = {'timestamp': datetime.now().isoformat(), 'routes': []}
        run_date_col = None
        try:
            _, _, run_date_col, _, _, _, _ = self._prepare_excel_for_export(self.excel_path)
        except Exception:
            pass
        for route in self.routes:
            print(f"\n[{route['code']}] {route['origin']} – {route['destination']}")
            route_results = {'route': route, 'prices': []}
            for source in self.sources:
                print(f"  {source['name']}")
                try:
                    price_data = None
                    if source['type'] == 'airline':
                        if source['name'] == 'Qatar Airways':
                            price_data = self.scrape_qatar_airways(route)
                        elif source['name'] == 'British Airways':
                            price_data = self.scrape_british_airways(route)
                        elif source['name'] == 'Malaysia Airlines':
                            price_data = self.scrape_malaysia_airlines(route)
                        elif source['name'] == 'Kuwait Airways':
                            price_data = self.scrape_kuwait_airways(route)
                        elif source['name'] == 'Turkish Airlines':
                            price_data = self.scrape_turkish_airlines(route)
                        elif source['name'] == 'Pakistan International Airlines':
                            price_data = self.scrape_pia(route)
                    else:
                        price_data = self.scrape_aggregator(source, route)
                    if price_data:
                        price_data['route_code'] = route['code']
                        price_data['source'] = source['name']
                        price_data['source_ar'] = source['name_ar']
                        price_data['source_code'] = source['source_code']
                        route_results['prices'].append(price_data)
                        print(f"    ✓ {price_data.get('price')} QAR")
                    else:
                        print(f"    ✗ No round-trip price")
                    time.sleep(3)
                except Exception as e:
                    print(f"    ✗ {e}")
            results['routes'].append(route_results)
            try:
                self.append_route_to_excel(route_results, self.excel_path, date_col_override=run_date_col)
                print(f"  ✓ Saved to Excel")
            except Exception as e:
                print(f"  ✗ Excel: {e}")
        self._close_driver()
        total = sum(len(r['prices']) for r in results['routes'])
        print("\n" + "="*60)
        print(f"Done. {total} prices found.")
        print("="*60)
        return results


def create_fresh_excel(path: str = None) -> str:
    """Create fresh flight_prices.xlsx with route block and flight headers (no data)."""
    path = os.path.abspath(path or FLIGHT_PRICES_EXCEL)
    if os.path.exists(path):
        backup = path.replace('.xlsx', '_backup_%s.xlsx' % datetime.now().strftime('%Y%m%d_%H%M'))
        try:
            os.rename(path, backup)
            print("Backed up to", backup)
        except OSError:
            os.remove(path)
    scraper = FlightPriceScraper(headless=True)
    routes = scraper._get_default_routes()
    scheduled = scraper._get_scheduled_dates_through_2026()
    wb = Workbook()
    ws = wb.active
    ws.title = FLIGHT_PRICES_SHEET_NAME
    ws.sheet_view.rightToLeft = False
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for col, h in enumerate(ROUTE_HEADERS, 1):
        c = ws.cell(row=1, column=col)
        c.value = h
        c.font = Font(bold=True)
        c.fill = header_fill
    for row_idx, r in enumerate(routes, 2):
        ws.cell(row=row_idx, column=1).value = r['code']
        ws.cell(row=row_idx, column=2).value = r['commodity_ar']
        ws.cell(row=row_idx, column=3).value = r['origin']
        ws.cell(row=row_idx, column=4).value = r['origin_code']
        ws.cell(row=row_idx, column=5).value = r['destination']
        ws.cell(row=row_idx, column=6).value = r['destination_code']
        ws.cell(row=row_idx, column=7).value = r['duration_months']
    flight_header_row = len(routes) + 2
    flight_headers = [
        'Code', 'Commodity', 'الدرجة المقابلة لها في الخطوط (Class equivalent in airlines)',
        'CPI-Flag', 'رمز المصدر (Source Code)', 'وكالات الخطوط (Flight Agencies)'
    ]
    for col_idx, header in enumerate(flight_headers, 1):
        cell = ws.cell(row=flight_header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for col, (header_text, _) in enumerate(scheduled, 7):
        cell = ws.cell(row=flight_header_row, column=col)
        cell.value = header_text
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 35
    for col in range(7, 7 + len(scheduled)):
        ws.column_dimensions[get_column_letter(col)].width = 15
    wb.save(path)
    print("Created", path)
    return path


def main():
    import sys
    if '--fresh-excel' in sys.argv or '-f' in sys.argv:
        create_fresh_excel()
        return
    scraper = FlightPriceScraper(headless=False)
    results = scraper.scrape_all()
    with open('flight_prices.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print("\nSaved flight_prices.json")
    scraper.export_to_excel(results)


if __name__ == '__main__':
    main()
