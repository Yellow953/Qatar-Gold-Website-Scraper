#!/usr/bin/env python3
"""Create a fresh flight_prices.xlsx (route block + headers, no data). Run: python3 create_fresh_flight_excel.py"""
import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FLIGHT_PRICES_EXCEL = 'flight_prices.xlsx'
FLIGHT_PRICES_SHEET_NAME = 'Flight Prices'
ROUTE_HEADERS = ['Code', 'Commodity', 'Origin', 'Origin_Code', 'Destination', 'Destination_Code', 'Duration_Months']
SCHEDULED_DAYS = (4, 10, 17, 24)

DEFAULT_ROUTES = [
    {'code': '007331101', 'commodity_ar': 'كلفة تذكرة دوحة _ لندن - دوحة لمدة 6 (Semi flexble التذكرة السياحية) أشهر', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'London', 'destination_code': 'LHR', 'duration_months': 6},
    {'code': '007331102', 'commodity_ar': 'كلفة تذكرة دوحة _ القاهرة - دوحة لمدة 6 (semi flexble التذكرة سياحية ( اشهر', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Cairo', 'destination_code': 'CAI', 'duration_months': 6},
    {'code': '007331103', 'commodity_ar': 'كلفة تذكرة دوحة_ كراتشي _ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Karachi', 'destination_code': 'KHI', 'duration_months': 6},
    {'code': '007331104', 'commodity_ar': 'كلفة تذكرة دوحة_ دبي _ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Dubai', 'destination_code': 'DXB', 'duration_months': 6},
    {'code': '007331105', 'commodity_ar': 'كلفة تذكرة دوحة_جدة _ دوحة لمدة 6 اشهر( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Jeddah', 'destination_code': 'JED', 'duration_months': 6},
    {'code': '007331106', 'commodity_ar': 'كلفة تذكرة دوحة_ بومباي _ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Mumbai', 'destination_code': 'BOM', 'duration_months': 6},
    {'code': '007331107', 'commodity_ar': 'كلفة تذكرة دوحة_كولا لمبور _ دوحة لمدة 6 اشهر( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Kuala Lumpur', 'destination_code': 'KUL', 'duration_months': 6},
    {'code': '007331108', 'commodity_ar': 'كلفة تذكرة دوحة_ اسطنبول لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Istanbul', 'destination_code': 'IST', 'duration_months': 6},
    {'code': '007331109', 'commodity_ar': 'كلفة تذكرة دوحة_ بانكوك _ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Bangkok', 'destination_code': 'BKK', 'duration_months': 6},
    {'code': '007331110', 'commodity_ar': 'كلفة تذكرة دوحة_تبليسي_ دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'Tbilisi', 'destination_code': 'TBS', 'duration_months': 6},
    {'code': '007331111', 'commodity_ar': 'كلفة تذكرة دوحة_نيويورك دوحة لمدة 6 اشهر ( التذكرة سياحية semi flexble)', 'origin': 'Doha', 'origin_code': 'DOH', 'destination': 'New York', 'destination_code': 'JFK', 'duration_months': 6},
]


def _scheduled_dates_through_2026():
    today = date.today()
    result = []
    for year in range(today.year, 2027):
        for month in range(1, 13):
            if year == 2026 and month > 12:
                break
            for day in SCHEDULED_DAYS:
                try:
                    d = date(year, month, day)
                    if d >= today:
                        result.append((d.strftime('%d-%b'),))
                except ValueError:
                    pass
    return result


def main():
    path = os.path.abspath(FLIGHT_PRICES_EXCEL)
    if os.path.exists(path):
        backup = path.replace('.xlsx', '_backup_%s.xlsx' % datetime.now().strftime('%Y%m%d_%H%M'))
        try:
            os.rename(path, backup)
            print('Backed up existing file to', backup)
        except OSError:
            os.remove(path)

    wb = Workbook()
    ws = wb.active
    ws.title = FLIGHT_PRICES_SHEET_NAME
    ws.sheet_view.rightToLeft = True
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col, h in enumerate(ROUTE_HEADERS, 1):
        c = ws.cell(row=1, column=col)
        c.value = h
        c.font = Font(bold=True)
        c.fill = header_fill
    for row_idx, r in enumerate(DEFAULT_ROUTES, 2):
        ws.cell(row=row_idx, column=1).value = r['code']
        ws.cell(row=row_idx, column=2).value = r['commodity_ar']
        ws.cell(row=row_idx, column=3).value = r['origin']
        ws.cell(row=row_idx, column=4).value = r['origin_code']
        ws.cell(row=row_idx, column=5).value = r['destination']
        ws.cell(row=row_idx, column=6).value = r['destination_code']
        ws.cell(row=row_idx, column=7).value = r['duration_months']

    flight_header_row = len(DEFAULT_ROUTES) + 2
    flight_headers = [
        'Code', 'Commodity', 'الدرجة المقابلة لها في الخطوط (Class equivalent in airlines)',
        'CPI-Flag', 'رمز المصدر (Source Code)', 'وكالات الخطوط (Flight Agencies)'
    ]
    for col_idx, header in enumerate(flight_headers, 1):
        cell = ws.cell(row=flight_header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    scheduled = _scheduled_dates_through_2026()
    for col, (header_text,) in enumerate(scheduled, 7):
        cell = ws.cell(row=flight_header_row, column=col)
        cell.value = header_text
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 35
    for col in range(7, 7 + len(scheduled)):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(path)
    print('Created fresh Excel:', path)


if __name__ == '__main__':
    main()
